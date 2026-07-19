"""Ingest a filled FP-adjudication workbook, measure the real FP rate, and
(optionally) switch off the matches confirmed wrong.

Companion to `build_fp_adjudication_xlsx.py` (roadmap B.4.1-4.2). Reads the
verdicts a reviewer entered, stores them in `match_adjudications` (so the
labels accumulate across review rounds), prints the measured false-positive
rate per matching method and similarity band, and with `--apply --commit`
supersedes the `unified_match_log` rows marked 'wrong'.

The measured rates are the deliverable: they replace the estimated per-method
FP rates with human-verified truth, which is what lets the matching thresholds
be tuned from evidence.

Invoke
------
  # Measure only (safe, no writes):
  py scripts/maintenance/ingest_fp_adjudication.py --in FP_Adjudication_2026-07-18.xlsx

  # Store labels (persist verdicts, still no match changes):
  py scripts/maintenance/ingest_fp_adjudication.py --in <file>.xlsx --store --commit

  # Store labels AND switch off confirmed-wrong matches:
  py scripts/maintenance/ingest_fp_adjudication.py --in <file>.xlsx --store --apply --commit
"""
from __future__ import annotations

import argparse
import os
import sys
from collections import defaultdict

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
from db_config import get_connection

VALID_VERDICTS = {"correct", "wrong", "unsure"}

DDL = """
CREATE TABLE IF NOT EXISTS match_adjudications (
    id              BIGSERIAL PRIMARY KEY,
    uml_id          BIGINT NOT NULL,
    match_method    TEXT,
    source_system   TEXT,
    source_name     TEXT,
    target_name     TEXT,
    similarity      NUMERIC,
    verdict         TEXT NOT NULL CHECK (verdict IN ('correct','wrong','unsure')),
    notes           TEXT,
    adjudicated_at  TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    UNIQUE (uml_id)
);
"""


def read_rows(path):
    """Read the 'Adjudicate' sheet -> list of dicts for rows with a verdict."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Adjudicate"] if "Adjudicate" in wb.sheetnames else wb.active
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}
    required = ("uml_id", "match_method" if "match_method" in idx else "method", "verdict")
    for col in ("uml_id", "verdict"):
        if col not in idx:
            raise SystemExit(f"Workbook missing required column '{col}'. Found: {header}")
    method_key = "method" if "method" in idx else "match_method"

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or idx["uml_id"] >= len(row):
            continue
        uml_id = row[idx["uml_id"]]
        verdict = row[idx["verdict"]]
        if uml_id is None or verdict is None:
            continue
        verdict = str(verdict).strip().lower()
        if verdict not in VALID_VERDICTS:
            continue
        out.append({
            "uml_id": int(uml_id),
            "method": row[idx[method_key]] if method_key in idx else None,
            "source_system": row[idx["source"]] if "source" in idx else None,
            "source_name": row[idx["source_name"]] if "source_name" in idx else None,
            "target_name": row[idx["matched_employer"]] if "matched_employer" in idx else None,
            "similarity": row[idx["similarity"]] if "similarity" in idx else None,
            "sim_band": row[idx["sim_band"]] if "sim_band" in idx else None,
            "verdict": verdict,
            "notes": row[idx["notes"]] if "notes" in idx else None,
        })
    return out


def measure(rows):
    """Print FP rate per method and per (method, band). wrong / (correct+wrong)."""
    by_method = defaultdict(lambda: defaultdict(int))
    by_band = defaultdict(lambda: defaultdict(int))
    for r in rows:
        by_method[r["method"]][r["verdict"]] += 1
        by_band[(r["method"], r["sim_band"])][r["verdict"]] += 1

    print("\n=== Measured false-positive rate per method ===")
    print(f"{'method':<26}{'correct':>8}{'wrong':>7}{'unsure':>8}{'FP rate':>10}")
    for m in sorted(by_method):
        d = by_method[m]
        c, w, u = d["correct"], d["wrong"], d["unsure"]
        denom = c + w
        rate = f"{100.0 * w / denom:.1f}%" if denom else "n/a"
        print(f"{str(m):<26}{c:>8}{w:>7}{u:>8}{rate:>10}")

    print("\n=== By similarity band ===")
    print(f"{'method':<26}{'band':<11}{'correct':>8}{'wrong':>7}{'FP rate':>10}")
    for key in sorted(by_band, key=lambda k: (str(k[0]), str(k[1]))):
        d = by_band[key]
        c, w = d["correct"], d["wrong"]
        denom = c + w
        rate = f"{100.0 * w / denom:.1f}%" if denom else "n/a"
        print(f"{str(key[0]):<26}{str(key[1]):<11}{c:>8}{w:>7}{rate:>10}")


def store(cur, rows):
    cur.execute(DDL)
    n = 0
    for r in rows:
        cur.execute(
            """
            INSERT INTO match_adjudications
                (uml_id, match_method, source_system, source_name, target_name,
                 similarity, verdict, notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (uml_id) DO UPDATE SET
                verdict = EXCLUDED.verdict,
                notes = EXCLUDED.notes,
                adjudicated_at = NOW()
            """,
            (r["uml_id"], r["method"], r["source_system"], r["source_name"],
             r["target_name"], r["similarity"], r["verdict"], r["notes"]),
        )
        n += 1
    return n


def apply_rejections(cur, rows):
    """Supersede active UML rows the reviewer marked 'wrong'. Reversible: status
    flip + evidence breadcrumb, not a delete."""
    wrong_ids = [r["uml_id"] for r in rows if r["verdict"] == "wrong"]
    if not wrong_ids:
        return 0
    cur.execute(
        """
        UPDATE unified_match_log
        SET status = 'rejected',
            evidence = COALESCE(evidence, '{}'::jsonb)
                || jsonb_build_object(
                    'rejected_reason', 'human_adjudication_fp',
                    'rejected_at', NOW()::text
                )
        WHERE id = ANY(%s) AND status = 'active'
        """,
        (wrong_ids,),
    )
    return cur.rowcount


def main():
    ap = argparse.ArgumentParser(description="Ingest FP adjudication verdicts")
    ap.add_argument("--in", dest="in_path", required=True, help="Filled .xlsx path")
    ap.add_argument("--store", action="store_true", help="Persist verdicts to match_adjudications")
    ap.add_argument("--apply", action="store_true", help="Supersede UML rows marked 'wrong'")
    ap.add_argument("--commit", action="store_true", help="Persist writes (else rollback)")
    args = ap.parse_args()

    rows = read_rows(args.in_path)
    print(f"Adjudicated rows read: {len(rows):,}")
    if not rows:
        print("No verdicts found (fill the 'verdict' column). Nothing to do.")
        return

    measure(rows)

    if not (args.store or args.apply):
        print("\n(measure-only; pass --store and/or --apply --commit to write)")
        return

    conn = get_connection()
    cur = conn.cursor()
    try:
        stored = store(cur, rows) if args.store else 0
        rejected = apply_rejections(cur, rows) if args.apply else 0
        if args.commit:
            conn.commit()
            print(f"\nCOMMITTED: stored {stored} labels; superseded {rejected} wrong matches.")
        else:
            conn.rollback()
            print(f"\nDRY-RUN: would store {stored} labels; would supersede {rejected} wrong "
                  f"matches. Re-run with --commit to persist.")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
