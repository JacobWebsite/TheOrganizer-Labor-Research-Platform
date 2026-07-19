"""Build an Excel workbook so a human can adjudicate matching false positives.

This is the "FP-rate adjudication tool" (roadmap B.4.1). It samples suspect
matches from `unified_match_log`, lays them out one-per-row with the source
name, the employer they were matched to, and a similarity score, and gives
each row a CORRECT / WRONG / UNSURE dropdown. A reviewer fills the verdict
column; `ingest_fp_adjudication.py` then reads it back, measures the true
false-positive rate per matching method, and (optionally) switches off the
matches confirmed wrong.

Why this exists
---------------
The platform's per-method FP rates are ESTIMATES from a text-similarity proxy
(R7 audit: PHONETIC ~90%, TRUNCATED ~73%, FUZZY ~80%, EIN ~60%). A false
positive means someone else's violations/contracts/revenue get stapled onto
an employer -- a credibility landmine on a dossier. This tool converts the
estimates into human-verified truth: which specific matches are wrong, and
what the real rate is per method, so the automatic thresholds can be set from
evidence instead of guesswork.

Sampling
--------
Per method, rows are bucketed into similarity bands (<0.30, 0.30-0.50,
0.50-0.70, >=0.70) and up to `--per-band` rows are drawn from each with a
fixed random seed. That spread lets the ingest measure the rate across the
whole similarity range (not just the worst tail) while still surfacing the
egregious low-similarity FPs for cleanup.

Invoke
------
  py scripts/maintenance/build_fp_adjudication_xlsx.py
  py scripts/maintenance/build_fp_adjudication_xlsx.py --per-band 15 \
      --methods EIN_EXACT,TRUNCATED_NAME_STATE,PHONETIC_STATE --out review.xlsx
"""
from __future__ import annotations

import argparse
import os
import random
import sys
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from psycopg2.extras import RealDictCursor

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
from db_config import get_connection

# High-FP methods worth a human's time (per R7/R8 audits + 2026-06-07 scoping).
# EIN_EXACT is limited to the unreliable-EIN sources where collisions happen.
DEFAULT_METHODS = [
    "EIN_EXACT",
    "TRUNCATED_NAME_STATE",
    "PHONETIC_STATE",
    "FUZZY_INMEMORY_TRIGRAM",
    "FUZZY_SPLINK_ADAPTIVE",
    "NAME_AGGRESSIVE_STATE",
]

# (label, low-inclusive, high-exclusive)
BANDS = [
    ("<0.30", -0.01, 0.30),
    ("0.30-0.50", 0.30, 0.50),
    ("0.50-0.70", 0.50, 0.70),
    (">=0.70", 0.70, 1.01),
]

HEADERS = [
    "row", "uml_id", "method", "source", "source_name",
    "matched_employer", "similarity", "sim_band", "verdict", "notes",
]


def _band_for(sim: float) -> str:
    for label, lo, hi in BANDS:
        if lo <= sim < hi:
            return label
    return ">=0.70"


def fetch_candidates(cur, methods):
    """Pull active matches for the given methods with source + target names."""
    cur.execute(
        """
        SELECT uml.id AS uml_id,
               uml.match_method,
               uml.source_system,
               uml.evidence->>'source_name' AS source_name,
               f.employer_name AS target_name,
               similarity(
                   LOWER(uml.evidence->>'source_name'),
                   LOWER(f.employer_name)
               ) AS sim
        FROM unified_match_log uml
        JOIN f7_employers_deduped f ON f.employer_id = uml.target_id
        WHERE uml.status = 'active'
          AND uml.match_method = ANY(%s)
          AND uml.evidence->>'source_name' IS NOT NULL
          AND uml.evidence->>'source_name' <> ''
        """,
        (methods,),
    )
    # For EIN_EXACT, only the unreliable-EIN sources are FP-prone; keep the
    # sample focused so reviewer time isn't spent on reliable OSHA/WHD EINs.
    rows = []
    for r in cur.fetchall():
        if r["match_method"] == "EIN_EXACT" and r["source_system"] not in ("990", "bmf", "mergent"):
            continue
        rows.append(r)
    return rows


def sample(rows, per_band, seed):
    """Stratified sample: up to `per_band` rows per (method, similarity band)."""
    rng = random.Random(seed)
    buckets = defaultdict(list)
    for r in rows:
        sim = float(r["sim"]) if r["sim"] is not None else 0.0
        buckets[(r["match_method"], _band_for(sim))].append(r)
    picked = []
    for key in sorted(buckets):
        pool = buckets[key]
        rng.shuffle(pool)
        picked.extend(pool[:per_band])
    # Order the workbook worst-first within each method so egregious FPs are
    # at the top, but keep methods grouped.
    picked.sort(key=lambda r: (r["match_method"], float(r["sim"] or 0.0)))
    return picked


def build_workbook(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Adjudicate"

    header_fill = PatternFill("solid", fgColor="2C2418")
    header_font = Font(bold=True, color="FAF6EF")
    input_fill = PatternFill("solid", fgColor="FFF3CD")

    # Header row
    for col, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for i, r in enumerate(rows, start=1):
        sim = round(float(r["sim"]) if r["sim"] is not None else 0.0, 3)
        excel_row = i + 1
        values = [
            i, r["uml_id"], r["match_method"], r["source_system"],
            r["source_name"], r["target_name"], sim, _band_for(sim), "", "",
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=v)
            if HEADERS[col - 1] in ("verdict", "notes"):
                cell.fill = input_fill

    # Dropdown on the verdict column (correct / wrong / unsure).
    verdict_col = get_column_letter(HEADERS.index("verdict") + 1)
    dv = DataValidation(
        type="list",
        formula1='"correct,wrong,unsure"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Pick correct, wrong, or unsure."
    dv.prompt = "Is this the SAME employer? correct / wrong / unsure"
    ws.add_data_validation(dv)
    if rows:
        dv.add(f"{verdict_col}2:{verdict_col}{len(rows) + 1}")

    widths = [5, 10, 24, 9, 40, 40, 10, 10, 12, 30]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    _add_instructions_sheet(wb, len(rows))
    wb.save(out_path)


def _add_instructions_sheet(wb, n_rows):
    ws = wb.create_sheet("How to use", 0)
    lines = [
        ("How to adjudicate matches", True),
        ("", False),
        (f"This workbook has {n_rows} matches to check on the 'Adjudicate' tab.", False),
        ("Each row is one link the system made between a government record", False),
        ("(the 'source_name') and an employer in our data ('matched_employer').", False),
        ("", False),
        ("Your job: decide if they are the SAME real employer.", True),
        ("  - In the yellow 'verdict' column, pick: correct, wrong, or unsure.", False),
        ("  - 'correct' = yes, same employer.", False),
        ("  - 'wrong'   = no, these are different employers (a false match).", False),
        ("  - 'unsure'  = can't tell from the names.", False),
        ("  - Use the yellow 'notes' column for anything worth remembering.", False),
        ("", False),
        ("The 'similarity' number is how alike the two names look (0-1).", False),
        ("Low numbers are usually wrong matches, but not always - trust the names.", False),
        ("", False),
        ("When done, save the file and tell Claude Code to ingest it:", True),
        ("  py scripts/maintenance/ingest_fp_adjudication.py --in <thisfile>.xlsx", False),
        ("That measures the real error rate per method and can switch off the", False),
        ("matches you marked 'wrong'.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=12 if i == 1 else 11)
    ws.column_dimensions["A"].width = 78


def main():
    ap = argparse.ArgumentParser(description="Build the FP adjudication workbook")
    ap.add_argument("--per-band", type=int, default=12,
                    help="Max rows per (method, similarity band). Default 12.")
    ap.add_argument("--methods", type=str, default=",".join(DEFAULT_METHODS),
                    help="Comma-separated match methods to sample.")
    ap.add_argument("--seed", type=int, default=42, help="Random seed for reproducibility.")
    ap.add_argument("--out", type=str, default=None, help="Output .xlsx path.")
    args = ap.parse_args()

    methods = [m.strip() for m in args.methods.split(",") if m.strip()]
    out_path = args.out or os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        "docs", "scratch", f"FP_Adjudication_{date.today().isoformat()}.xlsx",
    )

    conn = get_connection(cursor_factory=RealDictCursor)
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        candidates = fetch_candidates(cur, methods)
    finally:
        conn.close()

    rows = sample(candidates, args.per_band, args.seed)

    print(f"Candidate active matches (target methods): {len(candidates):,}")
    per_method = defaultdict(int)
    for r in rows:
        per_method[r["match_method"]] += 1
    print(f"Sampled into workbook: {len(rows):,}")
    for m in sorted(per_method):
        print(f"  {m}: {per_method[m]}")

    build_workbook(rows, out_path)
    print(f"\nWrote {out_path}")
    print("Fill the yellow 'verdict' column, then run ingest_fp_adjudication.py --in <file>.")


if __name__ == "__main__":
    main()
