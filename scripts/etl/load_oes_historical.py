#!/usr/bin/env python3
"""
load_oes_historical.py - Load OES (Occupational Employment Statistics) data 2003-2023
into PostgreSQL table oes_historical.

Source: BLS OES survey files from ~/Downloads/
  - 2003-2010: oesm{YY}in4.zip (national by 4-digit NAICS industry)
  - 2011-2023: oesm{YY}all.zip (all data, filtered to national industry-level)

Usage:
  python load_oes_historical.py --step all
  python load_oes_historical.py --step load --years 2003,2010,2020
  python load_oes_historical.py --step verify
"""

import argparse
import os
import sys
import tempfile
import zipfile
from io import StringIO
from pathlib import Path

from db_config import get_connection

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

DOWNLOADS = Path.home() / "Downloads"

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS oes_historical (
    year INTEGER NOT NULL,
    naics VARCHAR(8),
    naics_title VARCHAR(200),
    occ_code VARCHAR(10) NOT NULL,
    occ_title VARCHAR(200),
    o_group VARCHAR(10),
    tot_emp NUMERIC,
    emp_prse NUMERIC,
    pct_total NUMERIC,
    pct_rpt NUMERIC,
    h_mean NUMERIC,
    a_mean NUMERIC,
    mean_prse NUMERIC,
    h_median NUMERIC,
    a_median NUMERIC,
    h_pct10 NUMERIC, h_pct25 NUMERIC, h_pct75 NUMERIC, h_pct90 NUMERIC,
    a_pct10 NUMERIC, a_pct25 NUMERIC, a_pct75 NUMERIC, a_pct90 NUMERIC,
    _loaded_at TIMESTAMP DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_oesh_year ON oes_historical (year);
CREATE INDEX IF NOT EXISTS idx_oesh_naics ON oes_historical (naics);
CREATE INDEX IF NOT EXISTS idx_oesh_occ ON oes_historical (occ_code);
CREATE INDEX IF NOT EXISTS idx_oesh_year_naics_occ ON oes_historical (year, naics, occ_code);
"""

# Columns we want in the target table (excluding year and _loaded_at which are handled separately)
TARGET_COLS = [
    "naics", "naics_title", "occ_code", "occ_title", "o_group",
    "tot_emp", "emp_prse", "pct_total", "pct_rpt",
    "h_mean", "a_mean", "mean_prse",
    "h_median", "a_median",
    "h_pct10", "h_pct25", "h_pct75", "h_pct90",
    "a_pct10", "a_pct25", "a_pct75", "a_pct90",
]

NUMERIC_COLS = {
    "tot_emp", "emp_prse", "pct_total", "pct_rpt",
    "h_mean", "a_mean", "mean_prse",
    "h_median", "a_median",
    "h_pct10", "h_pct25", "h_pct75", "h_pct90",
    "a_pct10", "a_pct25", "a_pct75", "a_pct90",
}

# Non-numeric sentinel values from BLS (should become NULL)
NON_NUMERIC = {"*", "#", "**", "N/A", "n/a", "-", "N", "na", ""}

# ---------------------------------------------------------------------------
# Column mapping for early-year XLS files
# ---------------------------------------------------------------------------

# 2003-2010 xls files use GROUP instead of O_GROUP, and may vary slightly.
# We build a case-insensitive mapping from source column names to our target names.
COLUMN_ALIASES = {
    "naics": "naics",
    "naics_title": "naics_title",
    "occ_code": "occ_code",
    "occ code": "occ_code",       # 2014 and some other years use space
    "occ_title": "occ_title",
    "occ title": "occ_title",     # 2014 and some other years use space
    "group": "o_group",
    "o_group": "o_group",
    "i_group": "_i_group",        # industry group flag (2011+); not stored, used for filtering
    "own_code": "_own_code",      # ownership code (2011+); not stored
    "area": "_area",
    "area_title": "_area_title",
    "area_type": "_area_type",
    "prim_state": "_prim_state",
    "tot_emp": "tot_emp",
    "emp_prse": "emp_prse",
    "pct_total": "pct_total",
    "pct_rpt": "pct_rpt",
    "jobs_1000": "_jobs_1000",
    "loc_quotient": "_loc_quotient",
    "h_mean": "h_mean",
    "a_mean": "a_mean",
    "mean_prse": "mean_prse",
    "h_pct10": "h_pct10",
    "h_pct25": "h_pct25",
    "h_median": "h_median",
    "h_pct75": "h_pct75",
    "h_pct90": "h_pct90",
    "a_pct10": "a_pct10",
    "a_pct25": "a_pct25",
    "a_median": "a_median",
    "a_pct75": "a_pct75",
    "a_pct90": "a_pct90",
    "annual": "_annual",
    "hourly": "_hourly",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_numeric(val):
    """Convert a BLS value to a clean numeric string or empty string (for COPY NULL)."""
    if val is None:
        return ""
    s = str(val).strip()
    if s in NON_NUMERIC:
        return ""
    # Remove commas from numbers like "1,234,567"
    s = s.replace(",", "")
    # Some values have trailing % or other chars
    s = s.rstrip("%")
    try:
        float(s)
        return s
    except ValueError:
        return ""


def clean_text(val, maxlen=200):
    """Clean a text value for COPY."""
    if val is None:
        return ""
    s = str(val).strip()
    # Escape tabs and newlines for COPY format
    s = s.replace("\t", " ").replace("\n", " ").replace("\r", "")
    if maxlen:
        s = s[:maxlen]
    return s


def find_zip(year):
    """Find the OES zip file for a given year."""
    yy = f"{year % 100:02d}"
    if year <= 2010:
        # Industry-level zip
        fname = f"oesm{yy}in4.zip"
    else:
        fname = f"oesm{yy}all.zip"
    path = DOWNLOADS / fname
    if path.exists():
        return path
    return None


def find_xls_in_zip(zf, year):
    """Find the appropriate data file inside a zip archive.

    For 2003-2010 (in4 zips): look for nat4d*.xls
    For 2011-2023 (all zips): look for .xlsx file
    """
    names = zf.namelist()

    if year <= 2010:
        # Look for nat4d xls file (4-digit NAICS national data)
        candidates = [n for n in names if "nat4d" in n.lower() and n.lower().endswith(".xls")]
        if candidates:
            return candidates[0], "xls"
        # Fallback: any xls with "4d" or "4digit"
        candidates = [n for n in names if ("4d" in n.lower() or "4digit" in n.lower()) and n.lower().endswith(".xls")]
        if candidates:
            return candidates[0], "xls"
        # Last resort: first xls that has "nat" in name
        candidates = [n for n in names if "nat" in n.lower() and n.lower().endswith(".xls")]
        if candidates:
            return candidates[0], "xls"
        # Very last resort: first xls
        candidates = [n for n in names if n.lower().endswith(".xls") and not n.lower().endswith(".xlsx")]
        if candidates:
            return candidates[0], "xls"
        raise FileNotFoundError(f"No .xls file found in zip for {year}. Contents: {names}")
    else:
        # Look for .xlsx file
        candidates = [n for n in names if n.lower().endswith(".xlsx") and not n.startswith("__")]
        if candidates:
            # Prefer all_data or oes_data named files
            for c in candidates:
                bn = os.path.basename(c).lower()
                if "all_data" in bn or "oes_data" in bn or "oesm" in bn:
                    return c, "xlsx"
            return candidates[0], "xlsx"
        raise FileNotFoundError(f"No .xlsx file found in zip for {year}. Contents: {names}")


def read_xls_rows(zf, entry_name, year):
    """Read rows from a .xls file inside a zip. Returns (headers, rows)."""
    import xlrd

    data = zf.read(entry_name)
    # xlrd can read from bytes via file_contents
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheet_by_index(0)

    # First row is header
    raw_headers = [str(ws.cell_value(0, c)).strip().lower() for c in range(ws.ncols)]

    rows = []
    for r in range(1, ws.nrows):
        row = {}
        for c in range(ws.ncols):
            col_name = raw_headers[c]
            mapped = COLUMN_ALIASES.get(col_name)
            if mapped is None:
                continue
            val = ws.cell_value(r, c)
            # xlrd returns floats for numeric cells
            row[mapped] = val
        rows.append(row)

    return rows


def read_xlsx_rows(zf, entry_name, year):
    """Read rows from a .xlsx file inside a zip. Returns list of row dicts.

    Filters to AREA_TYPE=1 (national) and NAICS != '000000'.
    """
    from openpyxl import load_workbook

    # Extract xlsx to a temp file (openpyxl needs a file path for read_only mode)
    data = zf.read(entry_name)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(data)
        tmp_path = tmp.name

    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb.active

        # Read header row
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        raw_headers = [str(h).strip().lower() if h else "" for h in header_row]

        # Build column index mapping
        col_map = {}
        for idx, h in enumerate(raw_headers):
            mapped = COLUMN_ALIASES.get(h)
            if mapped:
                col_map[idx] = mapped

        # Find area_type and naics column indices for filtering
        area_type_idx = None
        naics_idx = None
        own_code_idx = None
        for idx, mapped in col_map.items():
            if mapped == "_area_type":
                area_type_idx = idx
            elif mapped == "naics":
                naics_idx = idx
            elif mapped == "_own_code":
                own_code_idx = idx

        rows = []
        row_count = 0
        skip_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1

            # Filter: national only (AREA_TYPE = 1)
            if area_type_idx is not None:
                area_type_val = str(row[area_type_idx]).strip() if row[area_type_idx] is not None else ""
                if area_type_val != "1":
                    skip_count += 1
                    continue

            # Filter: must have NAICS code (not cross-industry 000000)
            if naics_idx is not None:
                naics_val = str(row[naics_idx]).strip() if row[naics_idx] is not None else ""
                if naics_val in ("000000", "000000 ", "", "None"):
                    skip_count += 1
                    continue

            # Filter: only private + all ownership (OWN_CODE in 1235, or missing)
            # OWN_CODE: 1=Federal, 2=State, 3=Local, 5=Private, 1235=All
            # We want 1235 (cross-ownership) to get totals; if not available, accept all
            # Actually for industry-level national data, just keep all OWN_CODEs since
            # the primary use is industry wage/employment by occupation

            rec = {}
            for idx, mapped in col_map.items():
                if idx < len(row):
                    rec[mapped] = row[idx]
                else:
                    rec[mapped] = None

            rows.append(rec)

        wb.close()

        if row_count > 0:
            print(f"    xlsx: {row_count:,} total rows, {skip_count:,} filtered out, {len(rows):,} kept")

        return rows
    finally:
        os.unlink(tmp_path)


def build_copy_row(year, rec):
    """Build a tab-delimited row for COPY from a record dict."""
    parts = [str(year)]
    for col in TARGET_COLS:
        val = rec.get(col)
        if col in NUMERIC_COLS:
            parts.append(clean_numeric(val))
        else:
            parts.append(clean_text(val))
    return "\t".join(parts)


def get_available_years():
    """Scan Downloads directory for available OES zip files."""
    years = []
    for year in range(2003, 2024):
        if find_zip(year):
            years.append(year)
    return sorted(years)


# ---------------------------------------------------------------------------
# Steps
# ---------------------------------------------------------------------------

def step_schema(conn):
    """Create table and indexes."""
    print("=== SCHEMA ===")
    cur = conn.cursor()
    cur.execute(SCHEMA_SQL)
    conn.commit()
    print("  Table oes_historical and indexes created/verified.")


def step_load(conn, years):
    """Load data for specified years."""
    print(f"\n=== LOAD ({len(years)} years) ===")

    cur = conn.cursor()

    # Delete existing data for the years we're loading (idempotent per-year)
    for y in years:
        cur.execute("DELETE FROM oes_historical WHERE year = %s;", (y,))
    conn.commit()
    print(f"  Cleared oes_historical for {len(years)} years.")

    copy_cols = ["year"] + TARGET_COLS
    copy_sql = f"COPY oes_historical ({', '.join(copy_cols)}) FROM STDIN WITH (FORMAT text, NULL '')"

    total_loaded = 0

    for year in sorted(years):
        zip_path = find_zip(year)
        if not zip_path:
            print(f"  {year}: ZIP NOT FOUND, skipping")
            continue

        print(f"\n  {year}: {zip_path.name}")

        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                entry_name, fmt = find_xls_in_zip(zf, year)
                print(f"    Entry: {entry_name} ({fmt})")

                if fmt == "xls":
                    rows = read_xls_rows(zf, entry_name, year)
                else:
                    rows = read_xlsx_rows(zf, entry_name, year)

                if not rows:
                    print(f"    WARNING: No rows extracted for {year}")
                    continue

                # Build COPY buffer
                buf = StringIO()
                valid_count = 0
                for rec in rows:
                    # Require occ_code
                    occ = rec.get("occ_code")
                    if occ is None or str(occ).strip() == "":
                        continue
                    line = build_copy_row(year, rec)
                    buf.write(line + "\n")
                    valid_count += 1

                buf.seek(0)
                cur.copy_expert(copy_sql, buf)
                conn.commit()

                print(f"    Loaded {valid_count:,} rows")
                total_loaded += valid_count

        except Exception as e:
            print(f"    ERROR loading {year}: {e}")
            conn.rollback()
            raise

    print(f"\n  TOTAL: {total_loaded:,} rows loaded across {len(years)} years")
    return total_loaded


def step_verify(conn):
    """Verify loaded data."""
    print("\n=== VERIFY ===")
    cur = conn.cursor()

    # Total rows
    cur.execute("SELECT COUNT(*) FROM oes_historical;")
    total = cur.fetchone()[0]
    print(f"  Total rows: {total:,}")

    if total == 0:
        print("  No data loaded. Run --step load first.")
        return

    # Rows per year
    cur.execute("""
        SELECT year, COUNT(*), COUNT(DISTINCT naics), COUNT(DISTINCT occ_code)
        FROM oes_historical
        GROUP BY year ORDER BY year;
    """)
    print(f"\n  {'Year':<6} {'Rows':>10} {'NAICS':>8} {'Occs':>8}")
    print(f"  {'-'*6} {'-'*10} {'-'*8} {'-'*8}")
    for row in cur.fetchall():
        print(f"  {row[0]:<6} {row[1]:>10,} {row[2]:>8,} {row[3]:>8,}")

    # Sample data
    cur.execute("""
        SELECT year, naics, naics_title, occ_code, occ_title, tot_emp, a_mean
        FROM oes_historical
        WHERE tot_emp IS NOT NULL AND a_mean IS NOT NULL
        ORDER BY year, naics, occ_code
        LIMIT 5;
    """)
    print("\n  Sample rows:")
    print(f"  {'Year':<6} {'NAICS':<8} {'Industry':<30} {'OCC':<10} {'Occupation':<30} {'Emp':>12} {'AvgWage':>10}")
    for row in cur.fetchall():
        ind = str(row[2] or "")[:28]
        occ = str(row[4] or "")[:28]
        emp = f"{row[5]:,.0f}" if row[5] else "N/A"
        wage = f"${row[6]:,.0f}" if row[6] else "N/A"
        print(f"  {row[0]:<6} {row[1] or '':<8} {ind:<30} {row[3]:<10} {occ:<30} {emp:>12} {wage:>10}")

    # Wage range sanity check
    cur.execute("""
        SELECT year,
               ROUND(AVG(a_mean), 0) AS avg_wage,
               ROUND(MIN(a_mean), 0) AS min_wage,
               ROUND(MAX(a_mean), 0) AS max_wage
        FROM oes_historical
        WHERE a_mean IS NOT NULL AND a_mean > 0
        GROUP BY year ORDER BY year;
    """)
    print("\n  Wage sanity check (a_mean):")
    print(f"  {'Year':<6} {'AvgWage':>10} {'MinWage':>10} {'MaxWage':>12}")
    for row in cur.fetchall():
        print(f"  {row[0]:<6} ${row[1]:>9,} ${row[2]:>9,} ${row[3]:>11,}")

    # Coverage: how many years have data
    cur.execute("SELECT MIN(year), MAX(year), COUNT(DISTINCT year) FROM oes_historical;")
    mn, mx, ct = cur.fetchone()
    print(f"\n  Year range: {mn}-{mx}, {ct} distinct years")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Load OES historical data 2003-2023")
    parser.add_argument("--step", choices=["schema", "load", "verify", "all"],
                        default="all", help="Step to run (default: all)")
    parser.add_argument("--years", type=str, default=None,
                        help="Comma-separated years (e.g. 2003,2010,2020). Default: all found.")
    args = parser.parse_args()

    # Determine years
    if args.years:
        years = [int(y.strip()) for y in args.years.split(",")]
    else:
        years = get_available_years()

    if not years and args.step in ("load", "all"):
        print("No OES zip files found in Downloads. Expected oesm{YY}in4.zip or oesm{YY}all.zip")
        sys.exit(1)

    print("OES Historical Data Loader")
    print(f"  Step: {args.step}")
    if years:
        print(f"  Years: {years[0]}-{years[-1]} ({len(years)} files)")
    print(f"  Downloads: {DOWNLOADS}")

    conn = get_connection()

    try:
        if args.step in ("schema", "all"):
            step_schema(conn)

        if args.step in ("load", "all"):
            step_load(conn, years)

        if args.step in ("verify", "all"):
            step_verify(conn)

    finally:
        conn.close()

    print("\nDone.")


if __name__ == "__main__":
    main()
