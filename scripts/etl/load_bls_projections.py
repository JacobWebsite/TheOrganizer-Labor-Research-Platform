"""
BLS Industry & Occupation Projections ETL
==========================================
Refreshes bls_industry_projections, bls_occupation_projections,
bls_industry_occupation_matrix from fresh BLS downloads (2024-2034).
Creates bls_naics_sector_map reference table from industry-sectoring-plan.xlsx.

Usage:
    py scripts/etl/load_bls_projections.py                    # full reload
    py scripts/etl/load_bls_projections.py --dry-run           # parse only, no DB writes
    py scripts/etl/load_bls_projections.py --step industry     # single table
    py scripts/etl/load_bls_projections.py --step verify       # check counts
"""
import argparse
import csv
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from db_config import get_connection

import openpyxl
import psycopg2.extras
from glob import glob


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DEFAULT_DATA_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "BLS industry and occupation projections"
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def clean_numeric(val):
    """Convert Excel/CSV value to float or None."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', '--', '\u2014', '-', 'None'):  # em-dash variants
        return None
    s = s.replace(',', '')
    try:
        return float(s)
    except ValueError:
        return None


def clean_soc_code(raw):
    """Strip ='XX-XXXX' formula wrapper from CSV occupation codes."""
    if raw is None:
        return None
    s = str(raw).strip()
    s = s.replace('="', '').replace('"', '')
    return s if s else None


def classify_growth(change_pct):
    """Assign growth_category from employment_change_pct."""
    if change_pct is None:
        return 'unknown'
    if change_pct >= 7:
        return 'fastest_growing'
    elif change_pct >= 2:
        return 'growing'
    elif change_pct >= -2:
        return 'stable'
    elif change_pct >= -7:
        return 'declining'
    else:
        return 'fast_declining'


def derive_soc_level(soc_code, occ_type):
    """Derive SOC hierarchy level from code pattern."""
    if not soc_code:
        return None
    if soc_code == '00-0000':
        return 1  # total
    if occ_type and occ_type.lower().startswith('line'):
        return 5  # detailed
    # Summary levels by code pattern
    if soc_code.endswith('-0000'):
        return 2  # major group
    if soc_code.endswith('000'):
        return 3  # minor group
    if soc_code.endswith('00'):
        return 4  # broad occupation
    return 5  # detailed


def normalize_occ_type(raw):
    """Normalize 'Line item' -> 'Line Item', 'Summary' stays."""
    if not raw:
        return raw
    s = str(raw).strip()
    if s.lower() == 'line item':
        return 'Line Item'
    if s.lower() == 'summary':
        return 'Summary'
    return s


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------
def parse_industry_xlsx(data_dir):
    """Parse industry.xlsx Table 2.11 -> list of dicts."""
    path = os.path.join(data_dir, "industry.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Table 2.11']

    rows = []
    for i in range(3, ws.max_row + 1):  # row 1=title, 2=headers, 3+=data
        cells = [c.value for c in ws[i]]
        title = cells[0]
        code = cells[1]
        if not code or str(code).startswith('Source:') or str(code).startswith('Note:'):
            continue

        title_str = str(title).strip() if title else ''
        code_str = str(code).strip()
        itype = str(cells[2]).strip() if cells[2] else None

        emp_2024 = clean_numeric(cells[3])
        emp_2034 = clean_numeric(cells[4])
        emp_change = clean_numeric(cells[5])
        emp_change_pct = clean_numeric(cells[6])
        emp_cagr = clean_numeric(cells[7])
        out_2024 = clean_numeric(cells[8])
        out_2034 = clean_numeric(cells[9])
        out_cagr = clean_numeric(cells[10])

        rows.append({
            'matrix_code': code_str,
            'industry_title': title_str,
            'industry_type': itype,
            'employment_2024': emp_2024,
            'employment_2034': emp_2034,
            'employment_change': emp_change,
            'employment_change_pct': emp_change_pct,
            'employment_cagr': emp_cagr,
            'output_2024': out_2024,
            'output_2034': out_2034,
            'output_cagr': out_cagr,
            'growth_category': classify_growth(emp_change_pct),
            'display_level': 0,
        })

    wb.close()
    return rows


def parse_occupation_xlsx(data_dir):
    """Parse occupation (2).xlsx Table 1.2 -> list of dicts."""
    path = os.path.join(data_dir, "occupation (2).xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Table 1.2']

    rows = []
    for i in range(3, ws.max_row + 1):
        cells = [c.value for c in ws[i]]
        title = cells[0]
        code = cells[1]
        if not code:
            continue
        code_str = str(code).strip()
        # Skip footnote/source rows
        if code_str.startswith('Footnote') or code_str.startswith('Source') or code_str.startswith('Note') or code_str.startswith('['):
            continue
        # SOC codes match XX-XXXX pattern
        if not re.match(r'\d{2}-\d{4}', code_str):
            continue

        title_str = str(title).strip() if title else ''
        occ_type = normalize_occ_type(cells[2])

        emp_2024 = clean_numeric(cells[3])
        emp_2034 = clean_numeric(cells[4])
        emp_dist_2024 = clean_numeric(cells[5])
        emp_dist_2034 = clean_numeric(cells[6])
        emp_change = clean_numeric(cells[7])
        emp_change_pct = clean_numeric(cells[8])
        pct_self = clean_numeric(cells[9])
        openings = clean_numeric(cells[10])
        wage = clean_numeric(cells[11])
        education = str(cells[12]).strip() if cells[12] and str(cells[12]).strip() not in ('--', '\u2014', 'None') else None
        experience = str(cells[13]).strip() if cells[13] and str(cells[13]).strip() not in ('--', '\u2014', 'None') else None
        training = str(cells[14]).strip() if cells[14] and str(cells[14]).strip() not in ('--', '\u2014', 'None') else None

        soc_level = derive_soc_level(code_str, occ_type)
        growth_cat = classify_growth(emp_change_pct)

        rows.append({
            'soc_code': code_str,
            'occupation_title': title_str,
            'occupation_type': occ_type,
            'employment_2024': emp_2024,
            'employment_2034': emp_2034,
            'employment_dist_2024': emp_dist_2024,
            'employment_dist_2034': emp_dist_2034,
            'employment_change': emp_change,
            'employment_change_pct': emp_change_pct,
            'pct_self_employed': pct_self,
            'annual_openings': openings,
            'median_wage_2024': wage,
            'typical_education': education,
            'work_experience': experience,
            'on_job_training': training,
            'soc_level': soc_level,
            'growth_category': growth_cat,
        })

    wb.close()
    return rows


def parse_matrix_csvs(data_dir):
    """Parse all NEM CSV files -> list of tuples for batch insert."""
    pattern = os.path.join(data_dir, 'National Employment Matrix_IND_*.csv')
    csv_files = sorted(glob(pattern))

    rows = []
    skipped = 0
    errors = 0
    for filepath in csv_files:
        fname = os.path.basename(filepath)
        # Skip duplicate files (have "(1)" in name)
        if '(1)' in fname:
            skipped += 1
            continue

        # Extract industry code from filename
        m = re.search(r'National Employment Matrix_IND_(.+)\.csv', fname)
        if not m:
            continue
        industry_code = m.group(1)

        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    occ_code = clean_soc_code(row.get('Occupation Code'))
                    if not occ_code:
                        continue
                    occ_type = normalize_occ_type(row.get('Occupation Type'))

                    rows.append((
                        industry_code,
                        occ_code,
                        (row.get('Occupation Title') or '').strip(),
                        occ_type,
                        clean_numeric(row.get('2024 Employment')),
                        clean_numeric(row.get('2024 Percent of Industry')),
                        clean_numeric(row.get('2024 Percent of Occupation')),
                        clean_numeric(row.get('Projected 2034 Employment')),
                        clean_numeric(row.get('Projected 2034 Percent of Industry')),
                        clean_numeric(row.get('Projected 2034 Percent of Occupation')),
                        clean_numeric(row.get('Employment Change, 2024-2034')),
                        clean_numeric(row.get('Employment Percent Change, 2024-2034')),
                        clean_numeric(row.get('Display Level')),
                        clean_numeric(row.get('Occupation Sort')),
                    ))
        except Exception as e:
            errors += 1
            print(f"  ERROR reading {fname}: {e}")

    unique_industries = len(set(r[0] for r in rows))
    print(f"  CSV files found: {len(csv_files)}, skipped duplicates: {skipped}, errors: {errors}")
    print(f"  Unique industries: {unique_industries}, total rows: {len(rows)}")
    return rows


def parse_sectoring_plan(data_dir):
    """Parse industry-sectoring-plan.xlsx Stubs -> list of dicts."""
    path = os.path.join(data_dir, "industry-sectoring-plan.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Stubs']

    rows = []
    for i in range(2, ws.max_row + 1):  # row 1=header
        cells = [c.value for c in ws[i]]
        sector_num = cells[0]
        if sector_num is None:
            continue
        try:
            sector_num = int(sector_num)
        except (ValueError, TypeError):
            continue

        rows.append({
            'sector_number': sector_num,
            'bls_io_summary': str(cells[1]).strip() if cells[1] else None,
            'naics_2022': str(cells[2]).strip() if cells[2] else None,
            'sector_title': str(cells[3]).strip() if cells[3] else None,
        })

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------
def apply_schema_changes(conn):
    """Add new columns and tables if they don't exist."""
    cur = conn.cursor()

    # Add occupation_sort to matrix if missing
    cur.execute("""
        ALTER TABLE bls_industry_occupation_matrix
        ADD COLUMN IF NOT EXISTS occupation_sort INTEGER
    """)

    # Create sector map table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS bls_naics_sector_map (
            sector_number INTEGER PRIMARY KEY,
            bls_io_summary VARCHAR(20),
            naics_2022 VARCHAR(100),
            sector_title VARCHAR(300)
        )
    """)

    conn.commit()
    cur.close()
    print("  Schema changes applied")


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------
def load_industry(conn, rows):
    """TRUNCATE + batch INSERT bls_industry_projections."""
    cur = conn.cursor()
    cur.execute("TRUNCATE bls_industry_projections CASCADE")
    # Reset sequence
    cur.execute("SELECT setval(pg_get_serial_sequence('bls_industry_projections', 'id'), 1, false)")

    template = """INSERT INTO bls_industry_projections
        (matrix_code, industry_title, industry_type,
         employment_2024, employment_2034, employment_change,
         employment_change_pct, employment_cagr,
         output_2024, output_2034, output_cagr,
         growth_category, display_level)
        VALUES %s"""

    values = [(
        r['matrix_code'], r['industry_title'], r['industry_type'],
        r['employment_2024'], r['employment_2034'], r['employment_change'],
        r['employment_change_pct'], r['employment_cagr'],
        r['output_2024'], r['output_2034'], r['output_cagr'],
        r['growth_category'], r['display_level'],
    ) for r in rows]

    psycopg2.extras.execute_values(cur, template, values, page_size=200)
    conn.commit()
    cur.close()
    print(f"  Loaded {len(rows)} rows into bls_industry_projections")


def load_occupations(conn, rows):
    """TRUNCATE + batch INSERT bls_occupation_projections."""
    cur = conn.cursor()
    cur.execute("TRUNCATE bls_occupation_projections CASCADE")

    template = """INSERT INTO bls_occupation_projections
        (soc_code, occupation_title, occupation_type,
         employment_2024, employment_2034, employment_dist_2024, employment_dist_2034,
         employment_change, employment_change_pct,
         pct_self_employed, annual_openings, median_wage_2024,
         typical_education, work_experience, on_job_training,
         soc_level, growth_category)
        VALUES %s"""

    values = [(
        r['soc_code'], r['occupation_title'], r['occupation_type'],
        r['employment_2024'], r['employment_2034'], r['employment_dist_2024'], r['employment_dist_2034'],
        r['employment_change'], r['employment_change_pct'],
        r['pct_self_employed'], r['annual_openings'], r['median_wage_2024'],
        r['typical_education'], r['work_experience'], r['on_job_training'],
        r['soc_level'], r['growth_category'],
    ) for r in rows]

    psycopg2.extras.execute_values(cur, template, values, page_size=200)
    conn.commit()
    cur.close()
    print(f"  Loaded {len(rows)} rows into bls_occupation_projections")


def load_matrix(conn, rows):
    """TRUNCATE + batch INSERT bls_industry_occupation_matrix."""
    cur = conn.cursor()
    cur.execute("TRUNCATE bls_industry_occupation_matrix CASCADE")
    cur.execute("SELECT setval(pg_get_serial_sequence('bls_industry_occupation_matrix', 'id'), 1, false)")

    template = """INSERT INTO bls_industry_occupation_matrix
        (industry_code, occupation_code, occupation_title, occupation_type,
         emp_2024, emp_2024_pct_industry, emp_2024_pct_occupation,
         emp_2034, emp_2034_pct_industry, emp_2034_pct_occupation,
         emp_change, emp_change_pct, display_level, occupation_sort)
        VALUES %s"""

    psycopg2.extras.execute_values(cur, template, rows, page_size=2000)
    conn.commit()
    cur.close()
    print(f"  Loaded {len(rows)} rows into bls_industry_occupation_matrix")


def load_sector_map(conn, rows):
    """TRUNCATE + batch INSERT bls_naics_sector_map."""
    cur = conn.cursor()
    cur.execute("TRUNCATE bls_naics_sector_map CASCADE")

    template = """INSERT INTO bls_naics_sector_map
        (sector_number, bls_io_summary, naics_2022, sector_title)
        VALUES %s"""

    values = [(
        r['sector_number'], r['bls_io_summary'], r['naics_2022'], r['sector_title'],
    ) for r in rows]

    psycopg2.extras.execute_values(cur, template, values, page_size=200)
    conn.commit()
    cur.close()
    print(f"  Loaded {len(rows)} rows into bls_naics_sector_map")


# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------
def rebuild_views(conn):
    """Recreate views that depend on BLS tables."""
    cur = conn.cursor()

    # Drop views first to allow column changes
    for v in ['v_industry_top_occupations', 'v_industry_union_risk',
              'v_industry_growth_opportunities', 'v_high_wage_growing_occupations']:
        cur.execute(f"DROP VIEW IF EXISTS {v} CASCADE")

    # v_industry_top_occupations - most important for occupation drill-down
    cur.execute("""
        CREATE VIEW v_industry_top_occupations AS
        SELECT
            m.industry_code,
            p.industry_title,
            m.occupation_code,
            m.occupation_title,
            m.occupation_type,
            m.emp_2024,
            m.emp_2034,
            m.emp_change,
            m.emp_change_pct,
            m.emp_2024_pct_industry,
            m.display_level,
            m.occupation_sort,
            o.median_wage_2024,
            o.typical_education,
            o.annual_openings
        FROM bls_industry_occupation_matrix m
        LEFT JOIN bls_industry_projections p ON m.industry_code = p.matrix_code
        LEFT JOIN bls_occupation_projections o ON m.occupation_code = o.soc_code
        WHERE m.occupation_type = 'Line Item'
    """)

    # v_industry_union_risk - industry projections + union density
    cur.execute("""
        CREATE VIEW v_industry_union_risk AS
        SELECT
            p.matrix_code,
            p.industry_title,
            p.employment_2024,
            p.employment_2034,
            p.employment_change_pct,
            p.growth_category,
            d.union_density_pct
        FROM bls_industry_projections p
        LEFT JOIN v_naics_union_density d
            ON LEFT(p.matrix_code, 2) = d.naics_2digit
        WHERE p.industry_type = 'Summary'
          AND p.matrix_code NOT LIKE 'TE%%'
    """)

    # v_industry_growth_opportunities - growing industries with union density
    cur.execute("""
        CREATE VIEW v_industry_growth_opportunities AS
        SELECT
            p.matrix_code,
            p.industry_title,
            p.employment_change_pct,
            p.employment_2024,
            p.growth_category,
            d.union_density_pct
        FROM bls_industry_projections p
        LEFT JOIN v_naics_union_density d
            ON LEFT(p.matrix_code, 2) = d.naics_2digit
        WHERE p.employment_change_pct > 0
        ORDER BY p.employment_change_pct DESC
    """)

    # v_high_wage_growing_occupations
    cur.execute("""
        CREATE VIEW v_high_wage_growing_occupations AS
        SELECT
            soc_code,
            occupation_title,
            employment_2024,
            employment_2034,
            employment_change_pct,
            median_wage_2024,
            typical_education,
            annual_openings
        FROM bls_occupation_projections
        WHERE employment_change_pct > 5
          AND median_wage_2024 > 50000
          AND occupation_type = 'Line Item'
        ORDER BY employment_change_pct DESC
    """)

    conn.commit()
    cur.close()
    print("  Views rebuilt")


# ---------------------------------------------------------------------------
# Verify
# ---------------------------------------------------------------------------
def verify(conn):
    """Run data integrity checks after load."""
    cur = conn.cursor()
    ok = True

    # Row counts
    checks = [
        ('bls_industry_projections', 400, 500),
        ('bls_occupation_projections', 1000, 1200),
        ('bls_industry_occupation_matrix', 100000, 130000),
        ('bls_naics_sector_map', 200, 250),
    ]
    print("\n  Row counts:")
    for table, lo, hi in checks:
        cur.execute(f"SELECT COUNT(*) FROM {table}")
        cnt = cur.fetchone()[0]
        status = 'OK' if lo <= cnt <= hi else 'WARN'
        if status != 'OK':
            ok = False
        print(f"    {table}: {cnt:,} [{status}] (expected {lo}-{hi})")

    # Composite codes exist
    composites = ['31-330', '44-450', '48-490']
    cur.execute("SELECT matrix_code FROM bls_industry_projections WHERE matrix_code = ANY(%s)", [composites])
    found = [r[0] for r in cur.fetchall()]
    missing = [c for c in composites if c not in found]
    if missing:
        print(f"    WARN: Missing composite codes: {missing}")
        ok = False
    else:
        print(f"    Composite codes (31-330, 44-450, 48-490): OK")

    # NULL checks on matrix
    cur.execute("""
        SELECT
            COUNT(*) FILTER (WHERE industry_code IS NULL) as null_ind,
            COUNT(*) FILTER (WHERE occupation_code IS NULL) as null_occ
        FROM bls_industry_occupation_matrix
    """)
    r = cur.fetchone()
    if r[0] > 0 or r[1] > 0:
        print(f"    WARN: Matrix has {r[0]} NULL industry_codes, {r[1]} NULL occupation_codes")
        ok = False
    else:
        print(f"    Matrix NULL check: OK (0 NULLs)")

    # growth_category coverage
    cur.execute("SELECT COUNT(*) FILTER (WHERE growth_category IS NULL OR growth_category = 'unknown') FROM bls_industry_projections")
    unk = cur.fetchone()[0]
    print(f"    Growth category coverage: {unk} unknown/NULL")

    # Unique industry codes in matrix
    cur.execute("SELECT COUNT(DISTINCT industry_code) FROM bls_industry_occupation_matrix")
    print(f"    Unique industries in matrix: {cur.fetchone()[0]}")

    # Unique occupation codes in matrix
    cur.execute("SELECT COUNT(DISTINCT occupation_code) FROM bls_industry_occupation_matrix")
    print(f"    Unique occupations in matrix: {cur.fetchone()[0]}")

    # Views work
    for v in ['v_industry_top_occupations', 'v_high_wage_growing_occupations']:
        try:
            cur.execute(f"SELECT COUNT(*) FROM {v}")
            cnt = cur.fetchone()[0]
            print(f"    View {v}: {cnt:,} rows")
        except Exception as e:
            print(f"    WARN: View {v} failed: {e}")
            conn.rollback()
            ok = False

    cur.close()
    return ok


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Load BLS Industry & Occupation Projections")
    parser.add_argument('--dry-run', action='store_true', help='Parse only, no DB writes')
    parser.add_argument('--step', choices=['all', 'schema', 'industry', 'occupation', 'matrix', 'sectors', 'views', 'verify'],
                        default='all', help='Run specific step')
    parser.add_argument('--data-dir', default=DEFAULT_DATA_DIR, help='Path to BLS data directory')
    parser.add_argument('--verbose', action='store_true', help='Print detailed progress')
    args = parser.parse_args()

    print("=" * 60)
    print("BLS Industry & Occupation Projections ETL")
    print("=" * 60)
    print(f"Data dir: {args.data_dir}")
    print(f"Mode: {'DRY RUN' if args.dry_run else 'LIVE'}")
    print(f"Step: {args.step}")
    print()

    steps = [args.step] if args.step != 'all' else ['schema', 'industry', 'occupation', 'matrix', 'sectors', 'views', 'verify']

    # Parse data (always, even in single-step mode for verify)
    industry_rows = occ_rows = matrix_rows = sector_rows = None

    if 'industry' in steps or args.step == 'all':
        print("[1/7] Parsing industry.xlsx...")
        industry_rows = parse_industry_xlsx(args.data_dir)
        print(f"  Parsed {len(industry_rows)} industry rows")

    if 'occupation' in steps or args.step == 'all':
        print("[2/7] Parsing occupation (2).xlsx...")
        occ_rows = parse_occupation_xlsx(args.data_dir)
        print(f"  Parsed {len(occ_rows)} occupation rows")

    if 'matrix' in steps or args.step == 'all':
        print("[3/7] Parsing NEM CSV files...")
        matrix_rows = parse_matrix_csvs(args.data_dir)

    if 'sectors' in steps or args.step == 'all':
        print("[4/7] Parsing industry-sectoring-plan.xlsx...")
        sector_rows = parse_sectoring_plan(args.data_dir)
        print(f"  Parsed {len(sector_rows)} sector rows")

    if args.dry_run:
        print("\n" + "=" * 60)
        print("DRY RUN COMPLETE - no data loaded")
        print("=" * 60)
        if industry_rows:
            print(f"  Industry:    {len(industry_rows)} rows ready")
        if occ_rows:
            print(f"  Occupation:  {len(occ_rows)} rows ready")
        if matrix_rows:
            print(f"  Matrix:      {len(matrix_rows)} rows ready")
        if sector_rows:
            print(f"  Sectors:     {len(sector_rows)} rows ready")
        return

    # Connect and load
    conn = get_connection()

    if 'schema' in steps:
        print("[5/7] Applying schema changes...")
        apply_schema_changes(conn)

    if 'industry' in steps and industry_rows:
        print("[5/7] Loading industry projections...")
        load_industry(conn, industry_rows)

    if 'occupation' in steps and occ_rows:
        print("[5/7] Loading occupation projections...")
        load_occupations(conn, occ_rows)

    if 'matrix' in steps and matrix_rows:
        print("[6/7] Loading industry-occupation matrix...")
        load_matrix(conn, matrix_rows)

    if 'sectors' in steps and sector_rows:
        print("[6/7] Loading NAICS sector map...")
        load_sector_map(conn, sector_rows)

    if 'views' in steps:
        print("[7/7] Rebuilding views...")
        rebuild_views(conn)

    if 'verify' in steps:
        print("\n[VERIFY] Running integrity checks...")
        ok = verify(conn)
        if ok:
            print("\n  All checks PASSED")
        else:
            print("\n  Some checks had WARNINGS - review above")

    conn.close()

    print("\n" + "=" * 60)
    print("COMPLETE")
    print("=" * 60)


if __name__ == '__main__':
    main()
