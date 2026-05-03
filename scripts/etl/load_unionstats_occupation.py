#!/usr/bin/env python3
"""
ETL: Load UnionStats.com occupation-level union density data into PostgreSQL.

Source: occ_YYYY.xlsx files from UnionStats.com (Barry Hirsch & David Macpherson)
Target: unionstats_occupation table in olms_multiyear

Each file has a single sheet 'occ' with:
  Row 1: title row
  Row 2: blank
  Row 3: headers (COC, Occupation, Obs, Employment, Members, Covered, %Mem, %Cov)
  Row 4+: data (group headers have COC=None and ALL-CAPS occupation names)

Usage:
  python load_unionstats_occupation.py                  # runs all steps
  python load_unionstats_occupation.py --step schema    # create table only
  python load_unionstats_occupation.py --step load      # truncate + load data
  python load_unionstats_occupation.py --step verify    # show summary stats
"""

import argparse
import glob
import os
import re
import sys
from io import StringIO
from decimal import Decimal

import openpyxl
import psycopg2
import psycopg2.extras

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DATA_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "..", "data", "unionstats", "occupation",
)

DB_PARAMS = dict(
    dbname="olms_multiyear",
    user="postgres",
    password="Juniordog33!",
    host="localhost",
)

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS unionstats_occupation (
    year INTEGER NOT NULL,
    coc INTEGER,
    occupation TEXT NOT NULL,
    is_group_header BOOLEAN DEFAULT FALSE,
    obs INTEGER,
    employment_thousands NUMERIC,
    members_thousands NUMERIC,
    covered_thousands NUMERIC,
    pct_members NUMERIC,
    pct_covered NUMERIC,
    PRIMARY KEY (year, occupation)
);

CREATE INDEX IF NOT EXISTS idx_us_occ_year ON unionstats_occupation (year);
CREATE INDEX IF NOT EXISTS idx_us_occ_coc ON unionstats_occupation (coc);
CREATE INDEX IF NOT EXISTS idx_us_occ_group ON unionstats_occupation (is_group_header);
"""

COPY_COLUMNS = [
    "year", "coc", "occupation", "is_group_header", "obs",
    "employment_thousands", "members_thousands", "covered_thousands",
    "pct_members", "pct_covered",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def get_connection():
    return psycopg2.connect(**DB_PARAMS)


def is_mostly_uppercase(text: str) -> bool:
    """Return True if the alphabetic characters are predominantly uppercase.

    Group headers in UnionStats files use ALL CAPS occupation names like
    'MANAGEMENT, BUSINESS, AND FINANCIAL OCCUPATIONS'. We check that at
    least 85% of the alpha characters are uppercase, which avoids false
    positives on mixed-case detail rows.
    """
    alpha_chars = [c for c in text if c.isalpha()]
    if not alpha_chars:
        return False
    upper_count = sum(1 for c in alpha_chars if c.isupper())
    return upper_count / len(alpha_chars) >= 0.85


def extract_year_from_filename(filename: str) -> int | None:
    """Extract the 4-digit year from a filename like occ_2023.xlsx."""
    m = re.search(r"occ_(\d{4})\.xlsx$", filename)
    return int(m.group(1)) if m else None


def format_copy_value(val) -> str:
    """Format a Python value for the COPY text stream."""
    if val is None:
        return r"\N"
    if isinstance(val, bool):
        return "t" if val else "f"
    return str(val)


def parse_file(filepath: str, year: int) -> list[dict]:
    """Parse one occ_YYYY.xlsx file and return a list of row dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Try the expected sheet name, fall back to first sheet
    if "occ" in wb.sheetnames:
        ws = wb["occ"]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # Need at least 8 columns
        if len(row) < 8:
            continue

        coc_raw, occupation_raw, obs_raw, emp_raw, mem_raw, cov_raw, pct_mem_raw, pct_cov_raw = row[:8]

        # Skip completely blank rows
        if occupation_raw is None or str(occupation_raw).strip() == "":
            continue

        occupation = str(occupation_raw).strip()

        # Parse COC -- should be integer or None
        coc = None
        if coc_raw is not None:
            try:
                coc = int(float(coc_raw))
            except (ValueError, TypeError):
                coc = None

        # Detect group header: COC is None AND occupation is mostly uppercase
        is_group_header = (coc is None) and is_mostly_uppercase(occupation)

        # Parse numeric fields
        def safe_int(v):
            if v is None:
                return None
            try:
                return int(float(v))
            except (ValueError, TypeError):
                return None

        def safe_numeric(v):
            if v is None:
                return None
            try:
                return float(v)
            except (ValueError, TypeError):
                return None

        obs = safe_int(obs_raw)
        employment_thousands = safe_numeric(emp_raw)
        members_thousands = safe_numeric(mem_raw)
        covered_thousands = safe_numeric(cov_raw)
        pct_members = safe_numeric(pct_mem_raw)
        pct_covered = safe_numeric(pct_cov_raw)

        rows.append({
            "year": year,
            "coc": coc,
            "occupation": occupation,
            "is_group_header": is_group_header,
            "obs": obs,
            "employment_thousands": employment_thousands,
            "members_thousands": members_thousands,
            "covered_thousands": covered_thousands,
            "pct_members": pct_members,
            "pct_covered": pct_covered,
        })

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Steps
# ---------------------------------------------------------------------------
def step_schema():
    """Create the target table and indexes."""
    print("=== SCHEMA ===")
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(SCHEMA_SQL)
    conn.commit()
    cur.close()
    conn.close()
    print("Table unionstats_occupation created (or already exists).")


def step_load():
    """Truncate and reload all occ_YYYY.xlsx files."""
    print("=== LOAD ===")

    # Discover files
    pattern = os.path.join(DATA_DIR, "occ_*.xlsx")
    files = sorted(glob.glob(pattern))
    if not files:
        print(f"No occ_*.xlsx files found in {os.path.abspath(DATA_DIR)}")
        sys.exit(1)

    print(f"Found {len(files)} file(s) in {os.path.abspath(DATA_DIR)}")

    # Parse all files first
    all_rows = []
    for filepath in files:
        fname = os.path.basename(filepath)
        year = extract_year_from_filename(fname)
        if year is None:
            print(f"  SKIP {fname} (cannot extract year)")
            continue

        try:
            rows = parse_file(filepath, year)
            all_rows.extend(rows)
            n_groups = sum(1 for r in rows if r["is_group_header"])
            n_detail = len(rows) - n_groups
            print(f"  {fname}: {len(rows)} rows ({n_groups} groups, {n_detail} detail)")
        except Exception as e:
            print(f"  ERROR {fname}: {e}")
            continue

    if not all_rows:
        print("No data parsed. Exiting.")
        sys.exit(1)

    # Build COPY buffer
    buf = StringIO()
    for row in all_rows:
        line = "\t".join(format_copy_value(row[col]) for col in COPY_COLUMNS)
        buf.write(line + "\n")
    buf.seek(0)

    # Load into database
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("TRUNCATE TABLE unionstats_occupation;")
    print(f"\nTruncated unionstats_occupation.")

    cur.copy_from(buf, "unionstats_occupation", columns=COPY_COLUMNS, sep="\t", null=r"\N")
    conn.commit()

    cur.execute("SELECT COUNT(*) FROM unionstats_occupation;")
    count = cur.fetchone()[0]
    print(f"Loaded {count} rows into unionstats_occupation.")

    cur.close()
    conn.close()


def step_verify():
    """Show summary statistics."""
    print("=== VERIFY ===")
    conn = get_connection()
    cur = conn.cursor()

    # Check table exists
    cur.execute("""
        SELECT EXISTS (
            SELECT 1 FROM information_schema.tables
            WHERE table_name = 'unionstats_occupation'
        );
    """)
    if not cur.fetchone()[0]:
        print("Table unionstats_occupation does not exist. Run --step schema first.")
        cur.close()
        conn.close()
        return

    # Total rows
    cur.execute("SELECT COUNT(*) FROM unionstats_occupation;")
    total = cur.fetchone()[0]
    print(f"\nTotal rows: {total:,}")

    if total == 0:
        print("No data loaded yet.")
        cur.close()
        conn.close()
        return

    # Rows per year
    print("\nRows per year:")
    cur.execute("""
        SELECT year, COUNT(*) AS rows,
               SUM(CASE WHEN is_group_header THEN 1 ELSE 0 END) AS groups,
               SUM(CASE WHEN NOT is_group_header THEN 1 ELSE 0 END) AS detail
        FROM unionstats_occupation
        GROUP BY year
        ORDER BY year;
    """)
    print(f"  {'Year':<6} {'Total':>6} {'Groups':>7} {'Detail':>7}")
    print(f"  {'-'*6} {'-'*6} {'-'*7} {'-'*7}")
    for row in cur.fetchall():
        print(f"  {row[0]:<6} {row[1]:>6,} {row[2]:>7,} {row[3]:>7,}")

    # Year range
    cur.execute("SELECT MIN(year), MAX(year), COUNT(DISTINCT year) FROM unionstats_occupation;")
    min_yr, max_yr, n_years = cur.fetchone()
    print(f"\nYear range: {min_yr} - {max_yr} ({n_years} distinct years)")

    # Group headers vs detail
    cur.execute("""
        SELECT is_group_header, COUNT(*)
        FROM unionstats_occupation
        GROUP BY is_group_header
        ORDER BY is_group_header;
    """)
    for is_group, cnt in cur.fetchall():
        label = "Group headers" if is_group else "Detail rows"
        print(f"  {label}: {cnt:,}")

    # Sample group headers
    print("\nSample group headers (5 most recent):")
    cur.execute("""
        SELECT year, occupation, employment_thousands, pct_members
        FROM unionstats_occupation
        WHERE is_group_header
        ORDER BY year DESC, occupation
        LIMIT 5;
    """)
    for row in cur.fetchall():
        pct = f"{float(row[3]) * 100:.1f}%" if row[3] is not None else "N/A"
        emp = f"{float(row[2]):,.1f}k" if row[2] is not None else "N/A"
        print(f"  {row[0]} | {row[1][:60]:<60} | Emp: {emp:>10} | Mem: {pct}")

    # Sample detail rows
    print("\nSample detail occupations (5 most recent):")
    cur.execute("""
        SELECT year, coc, occupation, employment_thousands, pct_members
        FROM unionstats_occupation
        WHERE NOT is_group_header
        ORDER BY year DESC, occupation
        LIMIT 5;
    """)
    for row in cur.fetchall():
        pct = f"{float(row[4]) * 100:.1f}%" if row[4] is not None else "N/A"
        emp = f"{float(row[3]):,.1f}k" if row[3] is not None else "N/A"
        print(f"  {row[0]} | COC {row[1]:>4} | {row[2][:50]:<50} | Emp: {emp:>10} | Mem: {pct}")

    # Top union density occupations (most recent year)
    print(f"\nTop 10 highest union density occupations ({max_yr}):")
    cur.execute("""
        SELECT coc, occupation, pct_members, employment_thousands
        FROM unionstats_occupation
        WHERE year = %s AND NOT is_group_header AND pct_members IS NOT NULL
        ORDER BY pct_members DESC
        LIMIT 10;
    """, (max_yr,))
    for i, row in enumerate(cur.fetchall(), 1):
        pct = f"{float(row[2]) * 100:.1f}%"
        emp = f"{float(row[3]):,.1f}k" if row[3] is not None else "N/A"
        print(f"  {i:>2}. COC {row[0]:>4} | {row[1][:50]:<50} | {pct:>6} | Emp: {emp}")

    cur.close()
    conn.close()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Load UnionStats.com occupation-level union density data."
    )
    parser.add_argument(
        "--step",
        choices=["schema", "load", "verify", "all"],
        default="all",
        help="Which step to run (default: all)",
    )
    args = parser.parse_args()

    if args.step in ("schema", "all"):
        step_schema()
    if args.step in ("load", "all"):
        step_load()
    if args.step in ("verify", "all"):
        step_verify()

    print("\nDone.")


if __name__ == "__main__":
    main()
