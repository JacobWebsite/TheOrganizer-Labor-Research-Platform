"""
Load OES (Occupational Employment and Wage Statistics) data from BLS.

Source: Data_3_04/oesm24all.zip -> oesm24all/all_data_M_2024.xlsx
Table: oes_occupation_wages (~1.1-1.3M rows)
Curated MV: mv_oes_area_wages (cross-industry, detailed SOC only)

Usage:
  py scripts/etl/load_oes_wages.py
  py scripts/etl/load_oes_wages.py --data-dir Data_3_04 --dry-run
  py scripts/etl/load_oes_wages.py --step schema
  py scripts/etl/load_oes_wages.py --step load
  py scripts/etl/load_oes_wages.py --step views
  py scripts/etl/load_oes_wages.py --step verify
"""
from __future__ import annotations

import argparse
import os
import sys
import tempfile
import zipfile
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from db_config import get_connection


def ts():
    return datetime.now().strftime("%H:%M:%S")


CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS oes_occupation_wages (
    area VARCHAR(7),
    area_title VARCHAR(200),
    area_type INTEGER,
    prim_state VARCHAR(2),
    naics VARCHAR(8),
    naics_title VARCHAR(200),
    i_group VARCHAR(40),
    own_code INTEGER,
    occ_code VARCHAR(10) NOT NULL,
    occ_title VARCHAR(200),
    o_group VARCHAR(10),
    tot_emp INTEGER,
    emp_prse NUMERIC,
    jobs_1000 NUMERIC,
    loc_quotient NUMERIC,
    pct_total NUMERIC,
    h_mean NUMERIC,
    a_mean INTEGER,
    mean_prse NUMERIC,
    h_pct10 NUMERIC,
    h_pct25 NUMERIC,
    h_median NUMERIC,
    h_pct75 NUMERIC,
    h_pct90 NUMERIC,
    a_pct10 INTEGER,
    a_pct25 INTEGER,
    a_median INTEGER,
    a_pct75 INTEGER,
    a_pct90 INTEGER,
    annual BOOLEAN,
    hourly BOOLEAN,
    _loaded_at TIMESTAMP DEFAULT NOW()
);
"""

INDEX_SQLS = [
    "CREATE INDEX IF NOT EXISTS idx_oes_area ON oes_occupation_wages (area)",
    "CREATE INDEX IF NOT EXISTS idx_oes_occ ON oes_occupation_wages (occ_code)",
    "CREATE INDEX IF NOT EXISTS idx_oes_naics ON oes_occupation_wages (naics)",
    "CREATE INDEX IF NOT EXISTS idx_oes_area_occ ON oes_occupation_wages (area, occ_code)",
    "CREATE INDEX IF NOT EXISTS idx_oes_naics_occ ON oes_occupation_wages (naics, occ_code)",
    "CREATE INDEX IF NOT EXISTS idx_oes_state ON oes_occupation_wages (prim_state)",
]

MV_SQL = """
CREATE MATERIALIZED VIEW mv_oes_area_wages AS
SELECT area, area_title, area_type, prim_state, occ_code, occ_title,
       tot_emp, a_mean, a_median, a_pct10, a_pct25, a_pct75, a_pct90,
       h_median, loc_quotient
FROM oes_occupation_wages
WHERE i_group = 'cross-industry'
  AND o_group = 'detailed'
  AND area_type IN (2, 4);
"""

# Column name -> xlsx header mapping
COLS = [
    ("area", "AREA"),
    ("area_title", "AREA_TITLE"),
    ("area_type", "AREA_TYPE"),
    ("prim_state", "PRIM_STATE"),
    ("naics", "NAICS"),
    ("naics_title", "NAICS_TITLE"),
    ("i_group", "I_GROUP"),
    ("own_code", "OWN_CODE"),
    ("occ_code", "OCC_CODE"),
    ("occ_title", "OCC_TITLE"),
    ("o_group", "O_GROUP"),
    ("tot_emp", "TOT_EMP"),
    ("emp_prse", "EMP_PRSE"),
    ("jobs_1000", "JOBS_1000"),
    ("loc_quotient", "LOC_QUOTIENT"),
    ("pct_total", "PCT_TOTAL"),
    ("h_mean", "H_MEAN"),
    ("a_mean", "A_MEAN"),
    ("mean_prse", "MEAN_PRSE"),
    ("h_pct10", "H_PCT10"),
    ("h_pct25", "H_PCT25"),
    ("h_median", "H_MEDIAN"),
    ("h_pct75", "H_PCT75"),
    ("h_pct90", "H_PCT90"),
    ("a_pct10", "A_PCT10"),
    ("a_pct25", "A_PCT25"),
    ("a_median", "A_MEDIAN"),
    ("a_pct75", "A_PCT75"),
    ("a_pct90", "A_PCT90"),
    ("annual", "ANNUAL"),
    ("hourly", "HOURLY"),
]


def clean_suppressed(val):
    """Convert suppressed/special values to None, strip commas from numbers."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if s in ("*", "**", "#", "~", "", "-"):
        return None
    s = s.replace(",", "")
    return s


def parse_bool(val):
    if val is None:
        return None
    s = str(val).strip().lower()
    return s in ("true", "1", "yes")


def parse_int(val):
    v = clean_suppressed(val)
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def parse_float(val):
    v = clean_suppressed(val)
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_args():
    ap = argparse.ArgumentParser(description="Load OES occupation wage data")
    ap.add_argument("--data-dir", default="Data_3_04")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--step", choices=["schema", "load", "views", "verify"], default=None,
                    help="Run only one step (default: all)")
    return ap.parse_args()


def step_schema(conn):
    print(f"[{ts()}] Creating table...")
    with conn.cursor() as cur:
        cur.execute("DROP TABLE IF EXISTS oes_occupation_wages CASCADE")
        cur.execute(CREATE_TABLE_SQL)
        for idx_sql in INDEX_SQLS:
            cur.execute(idx_sql)
    conn.commit()
    print(f"[{ts()}] Schema ready")


def step_load(conn, data_dir: str, dry_run: bool):
    import openpyxl
    from psycopg2.extras import execute_values

    zip_path = os.path.join(data_dir, "oesm24all.zip")
    if not os.path.exists(zip_path):
        raise SystemExit(f"Missing: {zip_path}")

    print(f"[{ts()}] Extracting xlsx from {zip_path}...")
    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(zip_path) as zf:
            # Find the xlsx file inside
            xlsx_names = [n for n in zf.namelist() if n.endswith(".xlsx")]
            if not xlsx_names:
                raise SystemExit(f"No .xlsx found in {zip_path}. Contents: {zf.namelist()}")
            xlsx_name = xlsx_names[0]
            zf.extract(xlsx_name, tmpdir)
            xlsx_path = os.path.join(tmpdir, xlsx_name)

        print(f"[{ts()}] Loading workbook: {xlsx_name}")
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        # Build header map from first row
        header_map = {}
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for i, cell in enumerate(row):
                if cell:
                    header_map[str(cell).strip().upper()] = i

        # Verify required columns exist
        missing = [xlsx_col for _, xlsx_col in COLS if xlsx_col not in header_map]
        if missing:
            print(f"WARNING: missing columns in xlsx: {missing}")

        col_indices = []
        for db_col, xlsx_col in COLS:
            col_indices.append(header_map.get(xlsx_col))

        if dry_run:
            print(f"[{ts()}] DRY RUN - would load from {xlsx_name}")
            wb.close()
            return

        # Truncate
        with conn.cursor() as cur:
            cur.execute("TRUNCATE oes_occupation_wages")
        conn.commit()

        batch = []
        total = 0
        int_cols = {"area_type", "own_code", "tot_emp", "a_mean", "a_pct10", "a_pct25",
                     "a_median", "a_pct75", "a_pct90"}
        float_cols = {"emp_prse", "jobs_1000", "loc_quotient", "pct_total",
                       "h_mean", "mean_prse", "h_pct10", "h_pct25", "h_median",
                       "h_pct75", "h_pct90"}
        bool_cols = {"annual", "hourly"}

        print(f"[{ts()}] Reading rows...")
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values = []
            for j, (db_col, _) in enumerate(COLS):
                idx = col_indices[j]
                raw = row[idx] if idx is not None and idx < len(row) else None
                if db_col in int_cols:
                    values.append(parse_int(raw))
                elif db_col in float_cols:
                    values.append(parse_float(raw))
                elif db_col in bool_cols:
                    values.append(parse_bool(raw))
                else:
                    v = clean_suppressed(raw)
                    values.append(str(v) if v is not None else None)
            batch.append(tuple(values))

            if len(batch) >= 5000:
                with conn.cursor() as cur:
                    db_cols = ", ".join(c[0] for c in COLS)
                    execute_values(
                        cur,
                        f"INSERT INTO oes_occupation_wages ({db_cols}) VALUES %s",
                        batch,
                        page_size=5000,
                    )
                conn.commit()
                total += len(batch)
                batch = []
                if total % 100_000 == 0:
                    print(f"  [{ts()}] {total:,} rows loaded...")

        if batch:
            with conn.cursor() as cur:
                db_cols = ", ".join(c[0] for c in COLS)
                execute_values(
                    cur,
                    f"INSERT INTO oes_occupation_wages ({db_cols}) VALUES %s",
                    batch,
                    page_size=5000,
                )
            conn.commit()
            total += len(batch)

        wb.close()
        print(f"[{ts()}] Loaded {total:,} rows into oes_occupation_wages")


def step_views(conn):
    print(f"[{ts()}] Creating MV mv_oes_area_wages...")
    with conn.cursor() as cur:
        cur.execute("DROP MATERIALIZED VIEW IF EXISTS mv_oes_area_wages CASCADE")
        cur.execute(MV_SQL)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_mv_oes_area ON mv_oes_area_wages (area)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_mv_oes_occ ON mv_oes_area_wages (occ_code)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_mv_oes_state ON mv_oes_area_wages (prim_state)")
    conn.commit()

    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM mv_oes_area_wages")
        cnt = cur.fetchone()[0]
    print(f"[{ts()}] mv_oes_area_wages: {cnt:,} rows")


def step_verify(conn):
    print(f"\n[{ts()}] === Verification ===")
    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM oes_occupation_wages")
        total = cur.fetchone()[0]
        print(f"  oes_occupation_wages: {total:,} rows")

        # Software Developers (15-1252) national average
        cur.execute("""
            SELECT occ_code, occ_title, a_mean, a_median
            FROM oes_occupation_wages
            WHERE occ_code = '15-1252' AND area = '99' AND i_group = 'cross-industry'
            LIMIT 1
        """)
        row = cur.fetchone()
        if row:
            print(f"  Software Devs (15-1252) national: a_mean=${row[2]:,}, a_median=${row[3]:,}")
        else:
            # Try area code for national
            cur.execute("""
                SELECT occ_code, occ_title, a_mean, a_median
                FROM oes_occupation_wages
                WHERE occ_code = '15-1252' AND area_type = 1 AND i_group = 'cross-industry'
                LIMIT 1
            """)
            row = cur.fetchone()
            if row:
                print(f"  Software Devs (15-1252) national: a_mean=${row[2]:,}, a_median=${row[3]:,}")
            else:
                print("  WARNING: Could not find Software Devs national record")

        cur.execute("SELECT COUNT(*) FROM mv_oes_area_wages")
        mv_cnt = cur.fetchone()[0]
        print(f"  mv_oes_area_wages: {mv_cnt:,} rows")


def main():
    args = parse_args()
    # Resolve data dir relative to project root
    project_root = os.path.join(os.path.dirname(__file__), "..", "..")
    data_dir = os.path.join(project_root, args.data_dir)
    if not os.path.isdir(data_dir):
        data_dir = args.data_dir

    conn = get_connection()
    try:
        steps = [args.step] if args.step else ["schema", "load", "views", "verify"]
        for step in steps:
            if step in ("views", "verify") and args.dry_run:
                print(f"  DRY RUN: skipping {step}")
                continue
            if step == "schema":
                step_schema(conn)
            elif step == "load":
                step_load(conn, data_dir, args.dry_run)
            elif step == "views":
                step_views(conn)
            elif step == "verify":
                step_verify(conn)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
