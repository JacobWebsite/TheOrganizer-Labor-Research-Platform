#!/usr/bin/env python3
"""
load_bea_industry.py -- Load BEA industry-level data into PostgreSQL.

Datasets:
  1. BEA Value Added by Industry (ValueAdded.xlsx, sheet TVA105-A)
     -> bea_value_added table
  2. BEA KLEMS decomposition (KLEMS.xlsx, multiple sheets)
     -> bea_klems table
  3. BEA Fixed Assets by Industry (Section3All_xls.xlsx, sheet FAAt301ESI-A)
     -> bea_fixed_assets table (Table 3.1ESI: Current-Cost Net Stock)

Usage:
  python load_bea_industry.py                        # run all steps
  python load_bea_industry.py --step schema          # create tables only
  python load_bea_industry.py --step load-va         # load value added
  python load_bea_industry.py --step load-klems
  python load_bea_industry.py --step load-fixedassets
  python load_bea_industry.py --step verify
"""

import argparse
import sys
import io
import csv
import re
from pathlib import Path

import openpyxl
import psycopg2
import psycopg2.extras

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DB_PARAMS = dict(
    dbname="olms_multiyear",
    user="postgres",
    password="Juniordog33!",
    host="localhost",
)

VA_FILE = Path(r"C:\Users\jakew\Downloads\ValueAdded.xlsx")
KLEMS_FILE = Path(r"C:\Users\jakew\Downloads\KLEMS.xlsx")
FA_FILE = Path(r"C:\Users\jakew\Downloads\Section3All_xls.xlsx")

# Sheets to load from KLEMS.xlsx: code -> (sheet_name, title)
KLEMS_SHEETS = {
    "TKG105-A": ("TKG105-A", "Composition of Gross Output by Industry"),
    "TKG110-A": ("TKG110-A", "Shares of Gross Output by Industry"),
    "TKE103-A": ("TKE103-A", "Energy Inputs Quantity Index"),
    "TKM103-A": ("TKM103-A", "Materials Inputs Quantity Index"),
    "TKS103-A": ("TKS103-A", "Services Inputs Quantity Index"),
}

HEADER_ROW = 8  # 1-indexed row that contains column headers (Line, industry, blank, years...)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def get_conn():
    return psycopg2.connect(**DB_PARAMS)


def parse_indent(raw_name: str) -> int:
    """Count leading spaces and return indent level (2 spaces = 1 level)."""
    if not raw_name:
        return 0
    leading = len(raw_name) - len(raw_name.lstrip(" "))
    return leading // 2


def parse_sheet(wb, sheet_name: str):
    """
    Parse a BEA sheet with the standard layout:
      Row 8 header: Line | [industry] | [blank] | 1997 | 1998 | ...
      Rows 9+: data

    Returns (year_columns, rows) where:
      year_columns = list of int years
      rows = list of dicts with keys: line_number, industry, industry_clean,
             indent_level, values (dict year->numeric or None)
    """
    ws = wb[sheet_name]

    # --- Parse header row to find year columns ---
    header_cells = list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=False))[0]

    year_columns = []  # (col_index, year_int)
    for cell in header_cells:
        val = cell.value
        if val is None:
            continue
        # Year columns can be int or string like "1997"
        try:
            yr = int(val)
            if 1900 <= yr <= 2100:
                year_columns.append((cell.column - 1, yr))  # 0-indexed column
        except (ValueError, TypeError):
            pass

    if not year_columns:
        print(f"  WARNING: No year columns found in {sheet_name} header row {HEADER_ROW}")
        return [], []

    years = [yr for _, yr in year_columns]
    print(f"  Sheet {sheet_name}: {len(years)} year columns ({min(years)}-{max(years)})")

    # --- Parse data rows ---
    rows = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=False):
        cells = list(row)
        if not cells:
            continue

        # Column A = line number
        line_val = cells[0].value
        if line_val is None:
            continue
        try:
            line_number = int(line_val)
        except (ValueError, TypeError):
            continue

        # Column B = industry name (with leading spaces for hierarchy)
        raw_industry = cells[1].value
        if raw_industry is None or str(raw_industry).strip() == "":
            continue
        raw_industry = str(raw_industry)
        indent_level = parse_indent(raw_industry)
        industry_clean = raw_industry.strip()

        # Parse year values
        values = {}
        for col_idx, yr in year_columns:
            if col_idx < len(cells):
                cell_val = cells[col_idx].value
                if cell_val is not None and cell_val != "":
                    try:
                        values[yr] = float(cell_val)
                    except (ValueError, TypeError):
                        pass  # skip non-numeric

        rows.append(dict(
            line_number=line_number,
            industry=raw_industry,
            industry_clean=industry_clean,
            indent_level=indent_level,
            values=values,
        ))

    return years, rows


def copy_from_stringio(cur, table, columns, data_rows):
    """Bulk load using COPY via StringIO for speed."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter="\t", lineterminator="\n")
    for row in data_rows:
        writer.writerow(row)
    buf.seek(0)
    col_str = ", ".join(columns)
    cur.copy_expert(f"COPY {table} ({col_str}) FROM STDIN WITH (FORMAT csv, DELIMITER E'\\t', NULL '')", buf)
    return len(data_rows)


# ---------------------------------------------------------------------------
# Steps
# ---------------------------------------------------------------------------
def step_schema():
    """Create tables and indexes."""
    print("=== Creating schema ===")
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS bea_value_added (
            line_number INTEGER,
            industry TEXT NOT NULL,
            industry_clean TEXT NOT NULL,
            indent_level INTEGER,
            year INTEGER NOT NULL,
            value_millions NUMERIC,
            PRIMARY KEY (line_number, year)
        );
        CREATE INDEX IF NOT EXISTS idx_bea_va_year ON bea_value_added (year);
        CREATE INDEX IF NOT EXISTS idx_bea_va_industry ON bea_value_added (industry_clean);
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS bea_fixed_assets (
            line_number INTEGER,
            industry TEXT NOT NULL,
            industry_clean TEXT NOT NULL,
            indent_level INTEGER,
            year INTEGER NOT NULL,
            value_millions NUMERIC,
            PRIMARY KEY (line_number, year)
        );
        CREATE INDEX IF NOT EXISTS idx_bea_fa_year ON bea_fixed_assets (year);
        CREATE INDEX IF NOT EXISTS idx_bea_fa_industry ON bea_fixed_assets (industry_clean);
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS bea_klems (
            sheet_code VARCHAR(10) NOT NULL,
            sheet_title TEXT,
            line_number INTEGER,
            industry TEXT NOT NULL,
            industry_clean TEXT NOT NULL,
            indent_level INTEGER,
            year INTEGER NOT NULL,
            value NUMERIC,
            PRIMARY KEY (sheet_code, line_number, year)
        );
        CREATE INDEX IF NOT EXISTS idx_bea_klems_year ON bea_klems (year);
        CREATE INDEX IF NOT EXISTS idx_bea_klems_sheet ON bea_klems (sheet_code);
        CREATE INDEX IF NOT EXISTS idx_bea_klems_industry ON bea_klems (industry_clean);
    """)

    conn.commit()
    cur.close()
    conn.close()
    print("  Tables created: bea_value_added, bea_fixed_assets, bea_klems")


def step_load_va():
    """Load BEA Value Added by Industry."""
    print(f"=== Loading Value Added from {VA_FILE} ===")
    if not VA_FILE.exists():
        print(f"  ERROR: File not found: {VA_FILE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(VA_FILE, read_only=True, data_only=True)
    sheet_name = "TVA105-A"
    if sheet_name not in wb.sheetnames:
        print(f"  ERROR: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    years, rows = parse_sheet(wb, sheet_name)
    wb.close()

    if not rows:
        print("  ERROR: No data rows parsed.")
        sys.exit(1)

    # Flatten to (line_number, industry, industry_clean, indent_level, year, value_millions)
    flat_rows = []
    for r in rows:
        for yr, val in r["values"].items():
            flat_rows.append((
                r["line_number"],
                r["industry"],
                r["industry_clean"],
                r["indent_level"],
                yr,
                val,
            ))

    print(f"  Parsed {len(rows)} industries x {len(years)} years = {len(flat_rows)} rows")

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("TRUNCATE bea_value_added;")

    columns = ["line_number", "industry", "industry_clean", "indent_level", "year", "value_millions"]
    n = copy_from_stringio(cur, "bea_value_added", columns, flat_rows)

    conn.commit()
    cur.close()
    conn.close()
    print(f"  Loaded {n} rows into bea_value_added")


def step_load_fixedassets():
    """Load BEA Fixed Assets by Industry (Table 3.1ESI)."""
    print(f"=== Loading Fixed Assets from {FA_FILE} ===")
    if not FA_FILE.exists():
        print(f"  ERROR: File not found: {FA_FILE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(FA_FILE, read_only=True, data_only=True)
    sheet_name = "FAAt301ESI-A"
    if sheet_name not in wb.sheetnames:
        print(f"  ERROR: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    years, rows = parse_sheet(wb, sheet_name)
    wb.close()

    if not rows:
        print("  ERROR: No data rows parsed.")
        sys.exit(1)

    # Flatten to (line_number, industry, industry_clean, indent_level, year, value_millions)
    flat_rows = []
    for r in rows:
        for yr, val in r["values"].items():
            flat_rows.append((
                r["line_number"],
                r["industry"],
                r["industry_clean"],
                r["indent_level"],
                yr,
                val,
            ))

    print(f"  Parsed {len(rows)} industries x {len(years)} years = {len(flat_rows)} rows")

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("TRUNCATE bea_fixed_assets;")

    columns = ["line_number", "industry", "industry_clean", "indent_level", "year", "value_millions"]
    n = copy_from_stringio(cur, "bea_fixed_assets", columns, flat_rows)

    conn.commit()
    cur.close()
    conn.close()
    print(f"  Loaded {n} rows into bea_fixed_assets")


def step_load_klems():
    """Load BEA KLEMS sheets."""
    print(f"=== Loading KLEMS from {KLEMS_FILE} ===")
    if not KLEMS_FILE.exists():
        print(f"  ERROR: File not found: {KLEMS_FILE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(KLEMS_FILE, read_only=True, data_only=True)
    available = wb.sheetnames

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("TRUNCATE bea_klems;")

    total_loaded = 0
    for sheet_code, (sheet_name, sheet_title) in KLEMS_SHEETS.items():
        if sheet_name not in available:
            print(f"  WARNING: Sheet '{sheet_name}' not found. Available: {available}")
            continue

        years, rows = parse_sheet(wb, sheet_name)
        if not rows:
            print(f"  WARNING: No data rows in {sheet_name}")
            continue

        # Flatten
        flat_rows = []
        for r in rows:
            for yr, val in r["values"].items():
                flat_rows.append((
                    sheet_code,
                    sheet_title,
                    r["line_number"],
                    r["industry"],
                    r["industry_clean"],
                    r["indent_level"],
                    yr,
                    val,
                ))

        columns = ["sheet_code", "sheet_title", "line_number", "industry",
                    "industry_clean", "indent_level", "year", "value"]
        n = copy_from_stringio(cur, "bea_klems", columns, flat_rows)
        total_loaded += n
        print(f"  {sheet_code}: {len(rows)} industries, {len(flat_rows)} rows loaded")

    wb.close()
    conn.commit()
    cur.close()
    conn.close()
    print(f"  Total KLEMS rows loaded: {total_loaded}")


def step_verify():
    """Print row counts and samples."""
    print("=== Verification ===")
    conn = get_conn()
    cur = conn.cursor()

    # --- bea_value_added ---
    cur.execute("SELECT COUNT(*) FROM bea_value_added;")
    va_count = cur.fetchone()[0]
    print(f"\nbea_value_added: {va_count:,} rows")

    cur.execute("""
        SELECT COUNT(DISTINCT industry_clean) AS industries,
               COUNT(DISTINCT year) AS years,
               MIN(year), MAX(year)
        FROM bea_value_added;
    """)
    row = cur.fetchone()
    print(f"  {row[0]} industries, {row[1]} years ({row[2]}-{row[3]})")

    cur.execute("""
        SELECT industry_clean, indent_level, year, value_millions
        FROM bea_value_added
        WHERE year = (SELECT MAX(year) FROM bea_value_added)
        ORDER BY line_number
        LIMIT 10;
    """)
    print(f"\n  Sample (latest year):")
    for r in cur.fetchall():
        indent = "  " * r[1]
        val_str = f"${r[3]:,.0f}M" if r[3] is not None else "NULL"
        print(f"    {indent}{r[0]}: {val_str} ({r[2]})")

    # --- bea_fixed_assets ---
    cur.execute("SELECT COUNT(*) FROM bea_fixed_assets;")
    fa_count = cur.fetchone()[0]
    print(f"\nbea_fixed_assets: {fa_count:,} rows")

    cur.execute("""
        SELECT COUNT(DISTINCT industry_clean) AS industries,
               COUNT(DISTINCT year) AS years,
               MIN(year), MAX(year)
        FROM bea_fixed_assets;
    """)
    row = cur.fetchone()
    print(f"  {row[0]} industries, {row[1]} years ({row[2]}-{row[3]})")

    cur.execute("""
        SELECT industry_clean, indent_level, year, value_millions
        FROM bea_fixed_assets
        WHERE year = (SELECT MAX(year) FROM bea_fixed_assets)
        ORDER BY line_number
        LIMIT 10;
    """)
    print(f"\n  Sample (latest year):")
    for r in cur.fetchall():
        indent = "  " * r[1]
        val_str = f"${r[3]:,.0f}M" if r[3] is not None else "NULL"
        print(f"    {indent}{r[0]}: {val_str} ({r[2]})")

    # --- bea_klems ---
    cur.execute("SELECT COUNT(*) FROM bea_klems;")
    kl_count = cur.fetchone()[0]
    print(f"\nbea_klems: {kl_count:,} rows")

    cur.execute("""
        SELECT sheet_code, sheet_title, COUNT(*) AS rows,
               COUNT(DISTINCT industry_clean) AS industries,
               MIN(year), MAX(year)
        FROM bea_klems
        GROUP BY sheet_code, sheet_title
        ORDER BY sheet_code;
    """)
    print(f"\n  By sheet:")
    for r in cur.fetchall():
        print(f"    {r[0]} ({r[1]}): {r[2]:,} rows, {r[3]} industries, {r[4]}-{r[5]}")

    cur.execute("""
        SELECT sheet_code, industry_clean, year, value
        FROM bea_klems
        WHERE year = (SELECT MAX(year) FROM bea_klems)
        ORDER BY sheet_code, line_number
        LIMIT 10;
    """)
    print(f"\n  Sample (latest year):")
    for r in cur.fetchall():
        val_str = f"{r[3]}" if r[3] is not None else "NULL"
        print(f"    [{r[0]}] {r[1]}: {val_str} ({r[2]})")

    # --- Indent level distribution ---
    cur.execute("""
        SELECT indent_level, COUNT(DISTINCT industry_clean)
        FROM bea_value_added
        GROUP BY indent_level
        ORDER BY indent_level;
    """)
    print(f"\n  Value Added indent distribution:")
    for r in cur.fetchall():
        print(f"    Level {r[0]}: {r[1]} industries")

    cur.close()
    conn.close()
    print("\n=== Done ===")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Load BEA industry data into PostgreSQL")
    parser.add_argument(
        "--step",
        choices=["schema", "load-va", "load-klems", "load-fixedassets", "verify", "all"],
        default="all",
        help="Which step to run (default: all)",
    )
    args = parser.parse_args()

    steps = {
        "schema": step_schema,
        "load-va": step_load_va,
        "load-klems": step_load_klems,
        "load-fixedassets": step_load_fixedassets,
        "verify": step_verify,
    }

    if args.step == "all":
        for name, func in steps.items():
            func()
    else:
        # Schema is always needed before load steps
        if args.step in ("load-va", "load-klems", "load-fixedassets"):
            step_schema()
        steps[args.step]()


if __name__ == "__main__":
    main()
