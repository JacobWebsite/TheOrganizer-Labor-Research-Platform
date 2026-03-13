"""
Load Census RPE (Revenue Per Employee) ratios at national, state, and county level.

Sources:
  - EC2200BASIC.zip: National-level Economic Census (pipe-delimited)
  - us_state_6digitnaics_2022.xlsx: SUSB state x 6-digit NAICS
  - county_3digitnaics_2022.xlsx: SUSB county x 3-digit NAICS

Usage:
  py scripts/etl/load_census_rpe.py --file EC2200BASIC.zip
  py scripts/etl/load_census_rpe.py --susb-state us_state_6digitnaics_2022.xlsx
  py scripts/etl/load_census_rpe.py --susb-county county_3digitnaics_2022.xlsx
  py scripts/etl/load_census_rpe.py --status
"""

import argparse
import csv
import os
import sys
import zipfile
from pathlib import Path

import openpyxl
from psycopg2.extras import execute_values

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))
from db_config import get_connection

BATCH_SIZE = 5000

# FIPS state code -> 2-letter abbreviation
FIPS_TO_STATE = {
    '01': 'AL', '02': 'AK', '04': 'AZ', '05': 'AR', '06': 'CA',
    '08': 'CO', '09': 'CT', '10': 'DE', '11': 'DC', '12': 'FL',
    '13': 'GA', '15': 'HI', '16': 'ID', '17': 'IL', '18': 'IN',
    '19': 'IA', '20': 'KS', '21': 'KY', '22': 'LA', '23': 'ME',
    '24': 'MD', '25': 'MA', '26': 'MI', '27': 'MN', '28': 'MS',
    '29': 'MO', '30': 'MT', '31': 'NE', '32': 'NV', '33': 'NH',
    '34': 'NJ', '35': 'NM', '36': 'NY', '37': 'NC', '38': 'ND',
    '39': 'OH', '40': 'OK', '41': 'OR', '42': 'PA', '44': 'RI',
    '45': 'SC', '46': 'SD', '47': 'TN', '48': 'TX', '49': 'UT',
    '50': 'VT', '51': 'VA', '53': 'WA', '54': 'WV', '55': 'WI',
    '56': 'WY', '60': 'AS', '66': 'GU', '69': 'MP', '72': 'PR',
    '78': 'VI',
}


def create_table(conn, drop_existing=False):
    with conn.cursor() as cur:
        if drop_existing:
            cur.execute('DROP TABLE IF EXISTS census_rpe_ratios CASCADE')

        cur.execute("""
            CREATE TABLE IF NOT EXISTS census_rpe_ratios (
                id SERIAL PRIMARY KEY,
                naics_code TEXT NOT NULL,
                naics_title TEXT,
                revenue_total_thousands BIGINT,
                employee_count INTEGER,
                establishment_count INTEGER,
                annual_payroll_thousands BIGINT,
                rpe NUMERIC(14,2),
                pay_per_employee NUMERIC(14,2),
                year INTEGER DEFAULT 2022,
                loaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                state VARCHAR(2),
                county_fips VARCHAR(5),
                geo_level VARCHAR(10) DEFAULT 'national'
            )
        """)
        cur.execute('CREATE INDEX IF NOT EXISTS idx_rpe_naics ON census_rpe_ratios (naics_code)')
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_rpe_state_naics
            ON census_rpe_ratios (state, naics_code) WHERE state IS NOT NULL
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_rpe_county_naics
            ON census_rpe_ratios (county_fips, naics_code) WHERE county_fips IS NOT NULL
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_rpe_geo_level
            ON census_rpe_ratios (geo_level)
        """)
    conn.commit()


def ensure_geo_columns(conn):
    """Add geographic columns if they don't exist (for existing tables)."""
    with conn.cursor() as cur:
        for col, typ, default in [
            ('state', 'VARCHAR(2)', None),
            ('county_fips', 'VARCHAR(5)', None),
            ('geo_level', "VARCHAR(10)", "'national'"),
        ]:
            cur.execute("""
                SELECT 1 FROM information_schema.columns
                WHERE table_name = 'census_rpe_ratios' AND column_name = %s
            """, (col,))
            if not cur.fetchone():
                ddl = f"ALTER TABLE census_rpe_ratios ADD COLUMN {col} {typ}"
                if default:
                    ddl += f" DEFAULT {default}"
                cur.execute(ddl)
                print(f"  Added column {col}")

        # Backfill geo_level for existing national rows
        cur.execute("""
            UPDATE census_rpe_ratios SET geo_level = 'national'
            WHERE geo_level IS NULL
        """)

        # Create indexes
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_rpe_state_naics
            ON census_rpe_ratios (state, naics_code) WHERE state IS NOT NULL
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_rpe_county_naics
            ON census_rpe_ratios (county_fips, naics_code) WHERE county_fips IS NOT NULL
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_rpe_geo_level
            ON census_rpe_ratios (geo_level)
        """)
    conn.commit()


def parse_ec_data(zip_path):
    """Parse EC2200BASIC pipe-delimited data from zip.

    Filters to national-level rows (GEOTYPE='01') with valid EMP and RCPTOT.
    Returns list of tuples for insertion.
    """
    rows = []
    seen_naics = set()

    with zipfile.ZipFile(zip_path) as zf:
        dat_files = [n for n in zf.namelist() if n.endswith('.dat')]
        if not dat_files:
            print("No .dat file found in zip")
            return rows

        with zf.open(dat_files[0]) as f:
            reader = csv.DictReader(
                (line.decode('utf-8', errors='replace') for line in f),
                delimiter='|',
            )
            for row in reader:
                # Filter to US-wide totals only (GEOTYPE=01)
                if row.get('GEOTYPE', '') != '01':
                    continue

                # Filter to "All establishments" (TYPOP=00)
                if row.get('TYPOP', '') != '00':
                    continue

                # Filter to all tax statuses (TAXSTAT=00)
                if row.get('TAXSTAT', '') != '00':
                    continue

                naics = (row.get('NAICS2022') or '').strip()
                if not naics or naics == '00':
                    continue

                # Skip duplicate NAICS (can appear with different GEOCOMP)
                if naics in seen_naics:
                    continue

                emp_str = (row.get('EMP') or '').strip()
                rcptot_str = (row.get('RCPTOT') or '').strip()
                estab_str = (row.get('ESTAB') or '').strip()
                payann_str = (row.get('PAYANN') or '').strip()

                # Skip suppressed/missing data
                emp_flag = row.get('EMP_F', '').strip()
                if emp_flag and emp_flag != '0':
                    continue

                try:
                    emp = int(emp_str)
                    rcptot = int(rcptot_str)
                except (ValueError, TypeError):
                    continue

                if emp <= 0 or rcptot <= 0:
                    continue

                seen_naics.add(naics)

                estab = int(estab_str) if estab_str else None
                payann = int(payann_str) if payann_str else None
                naics_title = (row.get('NAICS2022_LABEL') or '').strip() or None

                rpe = round(rcptot * 1000 / emp, 2)  # Revenue in $1K -> dollars
                ppe = round(payann * 1000 / emp, 2) if payann else None

                rows.append((
                    naics, naics_title, rcptot, emp, estab,
                    payann, rpe, ppe,
                ))

    return rows


def load_rpe_data(conn, rows, drop_existing=False):
    """Load national RPE rows into census_rpe_ratios table."""
    create_table(conn, drop_existing)

    with conn.cursor() as cur:
        # Only truncate national rows (preserve state/county if already loaded)
        cur.execute("DELETE FROM census_rpe_ratios WHERE geo_level = 'national'")

        execute_values(cur, """
            INSERT INTO census_rpe_ratios
            (naics_code, naics_title, revenue_total_thousands, employee_count,
             establishment_count, annual_payroll_thousands, rpe, pay_per_employee)
            VALUES %s
        """, rows, page_size=BATCH_SIZE)

        cur.execute('SELECT COUNT(*) FROM census_rpe_ratios')
        cnt = cur.fetchone()[0]

    conn.commit()
    return cnt


def parse_susb_state(xlsx_path):
    """Parse SUSB state-level 6-digit NAICS data.

    Reads us_state_6digitnaics_2022.xlsx, filters to Total enterprise size,
    computes RPE per state+NAICS combination.

    Returns list of tuples:
      (naics, naics_title, receipts_1k, emp, estab, payroll_1k, rpe, ppe,
       state_abbrev, geo_level)
    """
    print(f"  Opening {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    seen = set()
    skipped_fips = 0
    skipped_filter = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:  # Skip title rows + header (rows 1-3)
            continue

        state_fips = str(row[0] or '').strip().zfill(2)
        naics = str(row[2] or '').strip()
        enterprise_size = str(row[4] or '').strip()

        # Skip national total (00) and aggregate NAICS (--)
        if state_fips == '00' or naics == '--':
            skipped_filter += 1
            continue

        # Only take Total rows
        if not enterprise_size.startswith('01:'):
            skipped_filter += 1
            continue

        # Map FIPS to state abbreviation
        state_abbrev = FIPS_TO_STATE.get(state_fips)
        if not state_abbrev:
            skipped_fips += 1
            continue

        # Parse numeric fields
        try:
            emp = int(row[7]) if row[7] is not None else 0
            receipts_1k = int(row[11]) if row[11] is not None else 0
        except (ValueError, TypeError):
            continue

        if emp <= 0 or receipts_1k <= 0:
            continue

        # Dedup by state+naics
        key = (state_abbrev, naics)
        if key in seen:
            continue
        seen.add(key)

        naics_title = str(row[3] or '').strip() or None
        try:
            estab = int(row[6]) if row[6] is not None else None
            payroll_1k = int(row[9]) if row[9] is not None else None
        except (ValueError, TypeError):
            estab = None
            payroll_1k = None

        rpe = round(receipts_1k * 1000 / emp, 2)
        ppe = round(payroll_1k * 1000 / emp, 2) if payroll_1k else None

        rows.append((
            naics, naics_title, receipts_1k, emp, estab,
            payroll_1k, rpe, ppe, state_abbrev, None, 'state',
        ))

    wb.close()

    if skipped_fips > 0:
        print(f"  Skipped {skipped_fips} rows with unknown FIPS codes")

    return rows


def parse_susb_county(xlsx_path):
    """Parse SUSB county-level 3-digit NAICS data.

    Reads county_3digitnaics_2022.xlsx, filters to Total enterprise size,
    computes RPE per county+NAICS combination.

    Returns list of tuples:
      (naics, naics_title, receipts_1k, emp, estab, payroll_1k, rpe, ppe,
       state_abbrev, county_fips_5char, geo_level)
    """
    print(f"  Opening {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    seen = set()
    skipped_fips = 0
    skipped_filter = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:  # Skip title rows + header (rows 1-3)
            continue

        state_fips = str(row[0] or '').strip().zfill(2)
        county_code = str(row[2] or '').strip().zfill(3)
        naics = str(row[4] or '').strip()
        enterprise_size = str(row[6] or '').strip()

        # Skip aggregate NAICS (--)
        if naics == '--':
            skipped_filter += 1
            continue

        # County file uses "1: Total" instead of "01: Total"
        if not enterprise_size.startswith('1:'):
            skipped_filter += 1
            continue

        # Map FIPS to state abbreviation
        state_abbrev = FIPS_TO_STATE.get(state_fips)
        if not state_abbrev:
            skipped_fips += 1
            continue

        county_fips = state_fips + county_code  # 5-char composite

        # Parse numeric fields
        try:
            emp = int(row[9]) if row[9] is not None else 0
            receipts_1k = int(row[13]) if row[13] is not None else 0
        except (ValueError, TypeError):
            continue

        if emp <= 0 or receipts_1k <= 0:
            continue

        # Dedup by county+naics
        key = (county_fips, naics)
        if key in seen:
            continue
        seen.add(key)

        naics_title = str(row[5] or '').strip() or None
        try:
            estab = int(row[8]) if row[8] is not None else None
            payroll_1k = int(row[11]) if row[11] is not None else None
        except (ValueError, TypeError):
            estab = None
            payroll_1k = None

        rpe = round(receipts_1k * 1000 / emp, 2)
        ppe = round(payroll_1k * 1000 / emp, 2) if payroll_1k else None

        rows.append((
            naics, naics_title, receipts_1k, emp, estab,
            payroll_1k, rpe, ppe, state_abbrev, county_fips, 'county',
        ))

    wb.close()

    if skipped_fips > 0:
        print(f"  Skipped {skipped_fips} rows with unknown FIPS codes")

    return rows


def load_geo_rpe_data(conn, rows, geo_level):
    """Load state or county RPE rows into census_rpe_ratios table."""
    ensure_geo_columns(conn)

    with conn.cursor() as cur:
        # Remove existing rows at this geo level
        cur.execute("DELETE FROM census_rpe_ratios WHERE geo_level = %s", (geo_level,))
        deleted = cur.rowcount
        if deleted > 0:
            print(f"  Deleted {deleted:,} existing {geo_level} rows")

        execute_values(cur, """
            INSERT INTO census_rpe_ratios
            (naics_code, naics_title, revenue_total_thousands, employee_count,
             establishment_count, annual_payroll_thousands, rpe, pay_per_employee,
             state, county_fips, geo_level)
            VALUES %s
        """, rows, page_size=BATCH_SIZE)

        cur.execute("SELECT COUNT(*) FROM census_rpe_ratios WHERE geo_level = %s",
                     (geo_level,))
        cnt = cur.fetchone()[0]

    conn.commit()
    return cnt


def show_status(conn):
    """Show RPE table stats."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT EXISTS(
                SELECT 1 FROM information_schema.tables
                WHERE table_name = 'census_rpe_ratios'
            ) AS e
        """)
        if not cur.fetchone()[0]:
            print("census_rpe_ratios table does not exist")
            return

        cur.execute("SELECT COUNT(*) FROM census_rpe_ratios")
        total = cur.fetchone()[0]
        print(f"census_rpe_ratios: {total:,} total rows")

        if total == 0:
            return

        # By geo level
        cur.execute("""
            SELECT COALESCE(geo_level, 'national') AS gl, COUNT(*) AS cnt
            FROM census_rpe_ratios
            GROUP BY COALESCE(geo_level, 'national')
            ORDER BY CASE COALESCE(geo_level, 'national')
                WHEN 'national' THEN 1 WHEN 'state' THEN 2 WHEN 'county' THEN 3
            END
        """)
        print("\n  By geographic level:")
        for row in cur.fetchall():
            print(f"    {row[0]:10s}: {row[1]:>10,}")

        # By NAICS granularity (national only)
        cur.execute("""
            SELECT LENGTH(naics_code) AS digits, COUNT(*) AS cnt
            FROM census_rpe_ratios
            WHERE geo_level = 'national' OR geo_level IS NULL
            GROUP BY LENGTH(naics_code)
            ORDER BY digits
        """)
        print("\n  National by NAICS granularity:")
        for row in cur.fetchall():
            print(f"    {row[0]}-digit: {row[1]:,}")

        # State coverage
        cur.execute("""
            SELECT COUNT(DISTINCT state) FROM census_rpe_ratios
            WHERE geo_level = 'state'
        """)
        state_cnt = cur.fetchone()[0]
        if state_cnt > 0:
            cur.execute("""
                SELECT COUNT(DISTINCT naics_code) FROM census_rpe_ratios
                WHERE geo_level = 'state'
            """)
            naics_cnt = cur.fetchone()[0]
            print(f"\n  State coverage: {state_cnt} states, {naics_cnt:,} unique NAICS codes")

        # County coverage
        cur.execute("""
            SELECT COUNT(DISTINCT county_fips) FROM census_rpe_ratios
            WHERE geo_level = 'county'
        """)
        county_cnt = cur.fetchone()[0]
        if county_cnt > 0:
            cur.execute("""
                SELECT COUNT(DISTINCT naics_code) FROM census_rpe_ratios
                WHERE geo_level = 'county'
            """)
            naics_cnt = cur.fetchone()[0]
            print(f"  County coverage: {county_cnt:,} counties, {naics_cnt:,} unique NAICS codes")

        # Top 10 RPE by 2-digit NAICS (national)
        cur.execute("""
            SELECT naics_code, naics_title, rpe, employee_count
            FROM census_rpe_ratios
            WHERE LENGTH(naics_code) = 2
              AND (geo_level = 'national' OR geo_level IS NULL)
            ORDER BY rpe DESC NULLS LAST
            LIMIT 10
        """)
        rows = cur.fetchall()
        if rows:
            print("\n  Top 10 RPE by 2-digit NAICS (national):")
            for row in rows:
                print(f"    {row[0]}  ${float(row[2]):>12,.0f}/emp  ({row[3]:>10,} emps)  {row[1]}")


def main():
    parser = argparse.ArgumentParser(description='Load Census RPE data (national + geographic)')
    parser.add_argument('--file', type=str, default=None,
                        help='Path to EC2200BASIC zip file (national)')
    parser.add_argument('--drop-existing', action='store_true',
                        help='Drop and recreate table')
    parser.add_argument('--susb-state', type=str, default=None,
                        help='Path to SUSB state-level xlsx (us_state_6digitnaics_2022.xlsx)')
    parser.add_argument('--susb-county', type=str, default=None,
                        help='Path to SUSB county-level xlsx (county_3digitnaics_2022.xlsx)')
    parser.add_argument('--status', action='store_true',
                        help='Show current status and exit')
    args = parser.parse_args()

    conn = get_connection()
    conn.autocommit = False

    if args.status:
        show_status(conn)
        conn.close()
        return

    # Handle national EC data
    if args.file:
        zip_path = args.file
        if not Path(zip_path).exists():
            print(f"File not found: {zip_path}")
            conn.close()
            return

        print(f"Parsing {zip_path}...")
        rows = parse_ec_data(zip_path)
        print(f"  Parsed {len(rows):,} NAICS codes with valid RPE data")

        if not rows:
            print("No data to load")
        else:
            cnt = load_rpe_data(conn, rows, args.drop_existing)
            print(f"  Loaded {cnt:,} total rows into census_rpe_ratios")
            print()

    # Handle SUSB state-level data
    if args.susb_state:
        xlsx_path = args.susb_state
        if not Path(xlsx_path).exists():
            print(f"File not found: {xlsx_path}")
            conn.close()
            return

        ensure_geo_columns(conn)
        print(f"Parsing SUSB state-level data from {xlsx_path}...")
        rows = parse_susb_state(xlsx_path)
        print(f"  Parsed {len(rows):,} state x NAICS combinations")

        if rows:
            cnt = load_geo_rpe_data(conn, rows, 'state')
            print(f"  Loaded {cnt:,} state-level rows")
        else:
            print("  No valid state data found")
        print()

    # Handle SUSB county-level data
    if args.susb_county:
        xlsx_path = args.susb_county
        if not Path(xlsx_path).exists():
            print(f"File not found: {xlsx_path}")
            conn.close()
            return

        ensure_geo_columns(conn)
        print(f"Parsing SUSB county-level data from {xlsx_path}...")
        rows = parse_susb_county(xlsx_path)
        print(f"  Parsed {len(rows):,} county x NAICS combinations")

        if rows:
            cnt = load_geo_rpe_data(conn, rows, 'county')
            print(f"  Loaded {cnt:,} county-level rows")
        else:
            print("  No valid county data found")
        print()

    # If no specific action, default to showing status
    if not args.file and not args.susb_state and not args.susb_county:
        print("No input file specified. Use --file, --susb-state, or --susb-county.")
        print("Showing current status:")
        print()

    show_status(conn)
    conn.close()


if __name__ == '__main__':
    main()
