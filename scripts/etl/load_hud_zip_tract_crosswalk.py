"""Load HUD USPS ZIP-Tract Crosswalk from Excel file.

Replaces the synthetic zip_tract_crosswalk (which had identical res_ratio
and bus_ratio for all rows) with official HUD data that has proper separate
residential and business allocation ratios.

Source: https://www.huduser.gov/portal/datasets/usps_crosswalk.html
  Select: ZIP-TRACT crosswalk, latest quarter

Usage:
    py scripts/etl/load_hud_zip_tract_crosswalk.py <path_to_xlsx>
    py scripts/etl/load_hud_zip_tract_crosswalk.py  # uses default path

Creates/replaces table: zip_tract_crosswalk
    zip_code    TEXT
    tract_geoid TEXT
    res_ratio   DOUBLE PRECISION  (residential allocation weight)
    bus_ratio   DOUBLE PRECISION  (business allocation weight)
"""
import sys
import os

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
from db_config import get_connection

DEFAULT_PATH = r"C:\Users\jakew\Downloads\ZIP_TRACT_122025.xlsx"


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH

    if not os.path.exists(xlsx_path):
        print(f"ERROR: File not found: {xlsx_path}")
        print("Download from: https://www.huduser.gov/portal/datasets/usps_crosswalk.html")
        sys.exit(1)

    print(f"Loading HUD ZIP-Tract crosswalk from: {xlsx_path}")

    # Read Excel file
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    # Verify header
    header = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    expected = ['ZIP', 'TRACT', 'USPS_ZIP_PREF_CITY', 'USPS_ZIP_PREF_STATE',
                'RES_RATIO', 'BUS_RATIO', 'OTH_RATIO', 'TOT_RATIO']
    if header != expected:
        print(f"ERROR: Unexpected header: {header}")
        print(f"Expected: {expected}")
        sys.exit(1)

    # Collect rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        zip_code = str(row[0]).zfill(5) if row[0] is not None else None
        tract_geoid = str(row[1]) if row[1] is not None else None
        res_ratio = float(row[4]) if row[4] is not None else 0.0
        bus_ratio = float(row[5]) if row[5] is not None else 0.0

        if zip_code and tract_geoid:
            rows.append((zip_code, tract_geoid, res_ratio, bus_ratio))

    wb.close()
    print(f"  Read {len(rows):,} rows from Excel")

    # Load into database
    conn = get_connection()
    cur = conn.cursor()

    # Get current stats for comparison
    cur.execute("SELECT EXISTS(SELECT 1 FROM pg_tables WHERE tablename = 'zip_tract_crosswalk')")
    table_exists = cur.fetchone()[0]
    old_count = 0
    if table_exists:
        cur.execute("SELECT COUNT(*) FROM zip_tract_crosswalk")
        old_count = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM zip_tract_crosswalk WHERE res_ratio != bus_ratio")
        old_diff = cur.fetchone()[0]
        print(f"  Old table: {old_count:,} rows, {old_diff:,} with different ratios")

    # Drop and recreate
    cur.execute("DROP TABLE IF EXISTS zip_tract_crosswalk CASCADE")
    cur.execute("""
        CREATE TABLE zip_tract_crosswalk (
            zip_code    TEXT NOT NULL,
            tract_geoid TEXT NOT NULL,
            res_ratio   DOUBLE PRECISION DEFAULT 0,
            bus_ratio   DOUBLE PRECISION DEFAULT 0
        )
    """)
    conn.commit()
    print("  Recreated table")

    # Bulk insert
    from io import StringIO
    buf = StringIO()
    for zip_code, tract_geoid, res_ratio, bus_ratio in rows:
        buf.write(f"{zip_code}\t{tract_geoid}\t{res_ratio}\t{bus_ratio}\n")
    buf.seek(0)
    cur.copy_from(buf, 'zip_tract_crosswalk',
                  columns=('zip_code', 'tract_geoid', 'res_ratio', 'bus_ratio'))
    conn.commit()
    print(f"  Inserted {len(rows):,} rows")

    # Create indexes
    cur.execute("CREATE INDEX idx_zip_tract_xwalk_zip ON zip_tract_crosswalk (zip_code)")
    cur.execute("CREATE INDEX idx_zip_tract_xwalk_tract ON zip_tract_crosswalk (tract_geoid)")
    conn.commit()
    print("  Created indexes")

    # Verify
    cur.execute("SELECT COUNT(*) FROM zip_tract_crosswalk")
    new_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM zip_tract_crosswalk WHERE res_ratio != bus_ratio")
    new_diff = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT zip_code) FROM zip_tract_crosswalk")
    n_zips = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT tract_geoid) FROM zip_tract_crosswalk")
    n_tracts = cur.fetchone()[0]

    print(f"\n  New table: {new_count:,} rows")
    print(f"  Rows where res_ratio != bus_ratio: {new_diff:,} ({new_diff*100/new_count:.1f}%)")
    print(f"  Coverage: {n_zips:,} ZIPs -> {n_tracts:,} tracts")
    if old_count:
        print(f"  Change: {old_count:,} -> {new_count:,} rows ({new_count - old_count:+,})")

    # Sample
    cur.execute("""
        SELECT zip_code, tract_geoid, res_ratio, bus_ratio
        FROM zip_tract_crosswalk
        WHERE res_ratio != bus_ratio
        LIMIT 5
    """)
    print("\n  Sample rows with different ratios:")
    for r in cur.fetchall():
        print(f"    ZIP {r[0]} -> Tract {r[1]}: res={r[2]:.4f}, bus={r[3]:.4f}")

    conn.close()
    print("\nDone.")


if __name__ == '__main__':
    main()
