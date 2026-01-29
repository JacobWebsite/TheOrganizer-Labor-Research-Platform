#!/usr/bin/env python3
"""
OLMS LM Filing Data Analysis Script
Analyzes LM-2, LM-3, LM-4 data for largest international unions

Usage:
    py olms_union_analysis_v2.py --data-dir "LM-2" --year 2023 --output union_report.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse
import os
import glob

# Excel formatting
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF')
MONEY_FORMAT = '$#,##0;($#,##0);"-"'
PCT_FORMAT = '0.0%'
NUM_FORMAT = '#,##0'

# Major international unions
MAJOR_UNIONS = {
    'SEIU': ['service employees international', 'seiu'],
    'Teamsters': ['teamsters', 'international brotherhood of teamsters', 'ibt'],
    'AFSCME': ['afscme', 'american federation of state county municipal', 'state county'],
    'UAW': ['united auto', 'uaw', 'automobile aerospace', 'united automobile'],
    'UFCW': ['ufcw', 'united food commercial workers', 'food and commercial'],
    'IBEW': ['ibew', 'electrical workers', 'international brotherhood of electrical'],
    'USW': ['steelworkers', 'usw', 'united steel', 'united steelworkers'],
    'CWA': ['cwa', 'communications workers'],
    'UNITE HERE': ['unite here', 'hotel employees restaurant'],
    'IAM': ['machinists', 'iam', 'international association of machinists'],
    'LIUNA': ['laborers', 'liuna', 'laborers international'],
    'Carpenters': ['carpenters', 'ubc', 'brotherhood of carpenters'],
    'AFT': ['aft', 'american federation of teachers'],
    'NEA': ['nea', 'national education association'],
    'IUOE': ['operating engineers', 'iuoe'],
    'Sheet Metal': ['sheet metal', 'smwia', 'smart'],
    'Plumbers': ['plumbers', 'pipefitters', 'ua local', 'united association'],
    'NALC': ['nalc', 'letter carriers'],
    'APWU': ['apwu', 'postal workers'],
    'IAFF': ['iaff', 'fire fighters', 'firefighters'],
    'Nurses': ['nurses', 'nnu', 'national nurses'],
    'OPEIU': ['opeiu', 'office professional employees'],
    'ATU': ['atu', 'transit union', 'amalgamated transit'],
    'BCTGM': ['bctgm', 'bakery confectionery'],
    'IATSE': ['iatse', 'theatrical stage'],
    'SAG-AFTRA': ['sag-aftra', 'screen actors', 'aftra'],
}


def load_olms_data(data_dir):
    """Load OLMS LM data files from extracted ZIP directory"""
    data = {}
    
    files = {
        'lm_data': 'lm_data',
        'assets_total': 'ar_assets_total',
        'liabilities_total': 'ar_liabilities_total',
        'receipts_total': 'ar_receipts_total',
        'disbursements_total': 'ar_disbursements_total',
        'officer_employee': 'ar_disbursements_emp_off',
        'membership': 'ar_membership',
        'payer_payee': 'ar_payer_payee',
        'disbursements_general': 'ar_disbursements_genrl',
    }
    
    for key, base_name in files.items():
        # Try different filename patterns
        patterns = [
            os.path.join(data_dir, f'{base_name}*.txt'),
            os.path.join(data_dir, f'{base_name.upper()}*.txt'),
            os.path.join(data_dir, f'*{base_name}*.txt'),
        ]
        
        filepath = None
        for pattern in patterns:
            matches = glob.glob(pattern)
            if matches:
                filepath = matches[0]
                break
        
        if filepath and os.path.exists(filepath):
            print(f"Loading {key}: {os.path.basename(filepath)}")
            try:
                df = pd.read_csv(
                    filepath, 
                    sep='|', 
                    low_memory=False,
                    encoding='latin-1',
                    on_bad_lines='skip'
                )
                data[key] = df
                print(f"  -> {len(df):,} records, {len(df.columns)} columns")
            except Exception as e:
                print(f"  ERROR: {e}")
                data[key] = pd.DataFrame()
        else:
            print(f"  File not found for {key}")
            data[key] = pd.DataFrame()
    
    return data


def classify_union(row):
    """Classify union to major international category"""
    check_fields = []
    for field in ['UNION_NAME', 'AFFIL_NAME', 'DESIG_NAME']:
        if field in row.index and pd.notna(row[field]):
            check_fields.append(str(row[field]).lower())
    
    combined_text = ' '.join(check_fields)
    
    for union, keywords in MAJOR_UNIONS.items():
        for keyword in keywords:
            if keyword in combined_text:
                return union
    
    union_indicators = ['local', 'district council', 'joint board', 'afl-cio', 'union']
    for indicator in union_indicators:
        if indicator in combined_text:
            return 'Other Union'
    
    return None


def analyze_unions(data):
    """Analyze union data and create summary"""
    
    if data['lm_data'].empty:
        print("ERROR: No main union data loaded")
        return None, None
    
    df = data['lm_data'].copy()
    
    # Map columns (case insensitive)
    col_map = {}
    for col in df.columns:
        col_lower = col.lower()
        if 'union_name' in col_lower:
            col_map['union_name'] = col
        elif 'f_num' in col_lower:
            col_map['f_num'] = col
        elif col_lower == 'members' or (col_lower.startswith('member') and 'ship' not in col_lower):
            col_map['members'] = col
        elif 'total_assets' in col_lower:
            col_map['total_assets'] = col
        elif 'total_receipts' in col_lower:
            col_map['total_receipts'] = col
        elif 'total_disb' in col_lower:
            col_map['total_disbursements'] = col
        elif 'affil' in col_lower and 'name' in col_lower:
            col_map['affil_name'] = col
        elif 'desig' in col_lower and 'name' in col_lower:
            col_map['desig_name'] = col
        elif 'form_type' in col_lower or col_lower == 'form_type':
            col_map['form_type'] = col
        elif col_lower == 'rpt_id':
            col_map['rpt_id'] = col
    
    print(f"\nMapped columns: {list(col_map.keys())}")
    
    # Rename columns
    df = df.rename(columns={v: k.upper() for k, v in col_map.items()})
    
    # Convert numeric
    for col in ['MEMBERS', 'TOTAL_ASSETS', 'TOTAL_RECEIPTS', 'TOTAL_DISBURSEMENTS']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Filter to LM-2
    if 'FORM_TYPE' in df.columns:
        df_lm2 = df[df['FORM_TYPE'].str.contains('LM-2', case=False, na=False)].copy()
        print(f"Filtered to LM-2: {len(df_lm2):,} records")
    else:
        df_lm2 = df.copy()
    
    # Classify unions
    print("Classifying unions...")
    df_lm2['UNION_CATEGORY'] = df_lm2.apply(classify_union, axis=1)
    
    df_unions = df_lm2[df_lm2['UNION_CATEGORY'].notna()].copy()
    print(f"Identified union records: {len(df_unions):,}")
    
    # Aggregate
    agg_cols = {}
    for col in ['MEMBERS', 'TOTAL_ASSETS', 'TOTAL_RECEIPTS', 'TOTAL_DISBURSEMENTS']:
        if col in df_unions.columns:
            agg_cols[col] = 'sum'
    
    if 'RPT_ID' in df_unions.columns:
        agg_cols['RPT_ID'] = 'count'
    elif 'UNION_NAME' in df_unions.columns:
        agg_cols['UNION_NAME'] = 'count'
    
    union_summary = df_unions.groupby('UNION_CATEGORY').agg(agg_cols).reset_index()
    
    # Rename count column
    for col in ['RPT_ID', 'UNION_NAME']:
        if col in union_summary.columns and col in agg_cols:
            union_summary = union_summary.rename(columns={col: 'FILING_COUNT'})
            break
    
    # Sort and calculate percentages
    if 'MEMBERS' in union_summary.columns:
        union_summary = union_summary.sort_values('MEMBERS', ascending=False)
        total = union_summary['MEMBERS'].sum()
        union_summary['PCT_OF_MEMBERS'] = union_summary['MEMBERS'] / total
        union_summary['CUMULATIVE_PCT'] = union_summary['PCT_OF_MEMBERS'].cumsum()
    
    return df_unions, union_summary


def analyze_officer_compensation(data, union_rpt_ids=None):
    """Analyze officer compensation"""
    
    if data['officer_employee'].empty:
        return pd.DataFrame()
    
    df = data['officer_employee'].copy()
    
    # Map columns
    col_map = {}
    for col in df.columns:
        col_lower = col.lower()
        if 'rpt_id' in col_lower:
            col_map[col] = 'RPT_ID'
        elif 'last_name' in col_lower or col_lower == 'lname':
            col_map[col] = 'LAST_NAME'
        elif 'first_name' in col_lower or col_lower == 'fname':
            col_map[col] = 'FIRST_NAME'
        elif 'title' in col_lower:
            col_map[col] = 'TITLE'
        elif 'gross_sal' in col_lower:
            col_map[col] = 'GROSS_SALARY'
        elif col_lower == 'total' or 'total_comp' in col_lower:
            col_map[col] = 'TOTAL_COMPENSATION'
        elif 'emp_off' in col_lower or 'empoff' in col_lower:
            col_map[col] = 'EMP_OFF_CODE'
    
    df = df.rename(columns=col_map)
    
    # Filter to unions
    if union_rpt_ids is not None and 'RPT_ID' in df.columns:
        df = df[df['RPT_ID'].isin(union_rpt_ids)]
    
    # Convert numeric
    for col in ['GROSS_SALARY', 'TOTAL_COMPENSATION']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Filter to officers (code 601)
    if 'EMP_OFF_CODE' in df.columns:
        df_off = df[df['EMP_OFF_CODE'] == 601]
        if len(df_off) > 0:
            df = df_off
    
    # Sort
    sort_col = 'TOTAL_COMPENSATION' if 'TOTAL_COMPENSATION' in df.columns else 'GROSS_SALARY'
    if sort_col in df.columns:
        df = df.sort_values(sort_col, ascending=False)
    
    return df


def create_excel_report(union_data, union_summary, officer_data, output_path, year=''):
    """Create Excel workbook"""
    
    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Executive Summary"
    
    rows = [
        ["OLMS Union Financial Analysis", ""],
        ["Data Year:", year],
        ["", ""],
        ["OVERVIEW", ""],
    ]
    
    if union_summary is not None:
        rows.append(["Total Union Categories:", len(union_summary)])
        if 'FILING_COUNT' in union_summary.columns:
            rows.append(["Total Filings:", int(union_summary['FILING_COUNT'].sum())])
        if 'MEMBERS' in union_summary.columns:
            rows.append(["Total Members:", f"{union_summary['MEMBERS'].sum():,.0f}"])
        if 'TOTAL_ASSETS' in union_summary.columns:
            rows.append(["Total Assets:", f"${union_summary['TOTAL_ASSETS'].sum():,.0f}"])
        if 'TOTAL_RECEIPTS' in union_summary.columns:
            rows.append(["Total Receipts:", f"${union_summary['TOTAL_RECEIPTS'].sum():,.0f}"])
        if 'TOTAL_DISBURSEMENTS' in union_summary.columns:
            rows.append(["Total Disbursements:", f"${union_summary['TOTAL_DISBURSEMENTS'].sum():,.0f}"])
    
    for row_idx, row_data in enumerate(rows, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx in [1, 4]:
                cell.font = Font(bold=True, size=14)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    
    # Sheet 2: Top Unions
    if union_summary is not None and not union_summary.empty:
        ws2 = wb.create_sheet("Top Unions by Members")
        headers = list(union_summary.columns)
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h.replace('_', ' '))
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        
        for row_idx, row in enumerate(union_summary.head(25).values, 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                h = headers[col_idx - 1]
                if any(x in h for x in ['ASSETS', 'RECEIPTS', 'DISBURSEMENTS']):
                    cell.number_format = MONEY_FORMAT
                elif 'PCT' in h:
                    cell.number_format = PCT_FORMAT
                elif any(x in h for x in ['MEMBERS', 'COUNT']):
                    cell.number_format = NUM_FORMAT
        
        for col_idx, h in enumerate(headers, 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = max(15, len(h) + 2)
    
    # Sheet 3: Officer Compensation
    if officer_data is not None and not officer_data.empty:
        ws3 = wb.create_sheet("Officer Compensation")
        cols = [c for c in ['LAST_NAME', 'FIRST_NAME', 'TITLE', 'GROSS_SALARY', 'TOTAL_COMPENSATION'] 
                if c in officer_data.columns]
        
        if cols:
            for col_idx, h in enumerate(cols, 1):
                cell = ws3.cell(row=1, column=col_idx, value=h.replace('_', ' '))
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            
            for row_idx, row in enumerate(officer_data[cols].head(500).values, 2):
                for col_idx, val in enumerate(row, 1):
                    cell = ws3.cell(row=row_idx, column=col_idx, value=val)
                    if cols[col_idx - 1] in ['GROSS_SALARY', 'TOTAL_COMPENSATION']:
                        cell.number_format = MONEY_FORMAT
            
            for col_idx in range(1, len(cols) + 1):
                ws3.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # Sheet 4: Union Details
    if union_data is not None and not union_data.empty:
        ws4 = wb.create_sheet("Union Details")
        cols = [c for c in ['UNION_CATEGORY', 'UNION_NAME', 'F_NUM', 'MEMBERS', 
                            'TOTAL_ASSETS', 'TOTAL_RECEIPTS', 'TOTAL_DISBURSEMENTS']
                if c in union_data.columns]
        
        if cols:
            sort_col = 'MEMBERS' if 'MEMBERS' in cols else cols[0]
            detail = union_data[cols].sort_values(sort_col, ascending=False).head(1000)
            
            for col_idx, h in enumerate(cols, 1):
                cell = ws4.cell(row=1, column=col_idx, value=h.replace('_', ' '))
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            
            for row_idx, row in enumerate(detail.values, 2):
                for col_idx, val in enumerate(row, 1):
                    cell = ws4.cell(row=row_idx, column=col_idx, value=val)
                    h = cols[col_idx - 1]
                    if any(x in h for x in ['ASSETS', 'RECEIPTS', 'DISBURSEMENTS']):
                        cell.number_format = MONEY_FORMAT
                    elif 'MEMBERS' in h:
                        cell.number_format = NUM_FORMAT
            
            for col_idx, h in enumerate(cols, 1):
                ws4.column_dimensions[get_column_letter(col_idx)].width = 20
    
    wb.save(output_path)
    print(f"\n*** Report saved to: {output_path} ***")


def main():
    parser = argparse.ArgumentParser(description='Analyze OLMS LM union data')
    parser.add_argument('--data-dir', required=True, help='Folder with OLMS data files')
    parser.add_argument('--year', default='2023', help='Year for report label')
    parser.add_argument('--output', default='union_report.xlsx', help='Output Excel file')
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("OLMS Union Financial Analysis")
    print("=" * 60)
    print(f"Data folder: {args.data_dir}")
    print(f"Output file: {args.output}")
    print("=" * 60)
    
    # Load
    print("\nLoading data files...")
    data = load_olms_data(args.data_dir)
    
    # Analyze
    print("\n" + "=" * 60)
    print("Analyzing unions...")
    union_data, union_summary = analyze_unions(data)
    
    if union_summary is not None and not union_summary.empty:
        print("\n" + "=" * 60)
        print("TOP 20 UNIONS BY MEMBERS:")
        print("=" * 60)
        print(union_summary.head(20).to_string(index=False))
    
    # Officer comp
    print("\n" + "=" * 60)
    print("Analyzing officer compensation...")
    rpt_ids = None
    if union_data is not None and 'RPT_ID' in union_data.columns:
        rpt_ids = union_data['RPT_ID'].unique()
    officer_data = analyze_officer_compensation(data, rpt_ids)
    print(f"Found {len(officer_data):,} officer/employee records")
    
    # Create report
    print("\n" + "=" * 60)
    print("Creating Excel report...")
    create_excel_report(union_data, union_summary, officer_data, args.output, args.year)
    
    print("\n" + "=" * 60)
    print("DONE!")
    print("=" * 60)


if __name__ == "__main__":
    main()
