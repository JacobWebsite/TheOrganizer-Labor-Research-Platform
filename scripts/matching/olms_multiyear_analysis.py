#!/usr/bin/env python3
"""
OLMS Multi-Year Union Financial Analysis (2000-2025)
Analyzes LM-2 data across multiple years for trend analysis

Usage:
    py olms_multiyear_analysis.py --data-dir "LM-2" --output union_trends.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse
import os
import glob

# Excel formatting
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF')
MONEY_FORMAT = '$#,##0;($#,##0);"-"'
PCT_FORMAT = '0.0%'
NUM_FORMAT = '#,##0'

# Major international unions
MAJOR_UNIONS = {
    'SEIU': ['service employees international', 'seiu'],
    'Teamsters': ['teamsters', 'international brotherhood of teamsters', 'ibt'],
    'AFSCME': ['afscme', 'american federation of state county municipal', 'state county'],
    'UAW': ['united auto', 'uaw', 'automobile aerospace', 'united automobile'],
    'UFCW': ['ufcw', 'united food commercial workers', 'food and commercial'],
    'IBEW': ['ibew', 'electrical workers', 'international brotherhood of electrical'],
    'USW': ['steelworkers', 'usw', 'united steel', 'united steelworkers'],
    'CWA': ['cwa', 'communications workers'],
    'UNITE HERE': ['unite here', 'hotel employees restaurant'],
    'IAM': ['machinists', 'iam', 'international association of machinists'],
    'LIUNA': ['laborers', 'liuna', 'laborers international'],
    'Carpenters': ['carpenters', 'ubc', 'brotherhood of carpenters'],
    'AFT': ['aft', 'american federation of teachers'],
    'NEA': ['nea', 'national education association'],
    'IUOE': ['operating engineers', 'iuoe'],
    'Sheet Metal': ['sheet metal', 'smwia', 'smart'],
    'Plumbers': ['plumbers', 'pipefitters', 'ua local', 'united association'],
    'NALC': ['nalc', 'letter carriers'],
    'APWU': ['apwu', 'postal workers'],
    'IAFF': ['iaff', 'fire fighters', 'firefighters'],
    'Nurses': ['nurses', 'nnu', 'national nurses'],
    'OPEIU': ['opeiu', 'office professional employees'],
    'ATU': ['atu', 'transit union', 'amalgamated transit'],
    'BCTGM': ['bctgm', 'bakery confectionery'],
    'IATSE': ['iatse', 'theatrical stage'],
    'SAG-AFTRA': ['sag-aftra', 'screen actors', 'aftra'],
}


def find_year_folders(data_dir):
    """Find all year folders in the data directory"""
    years = []
    for item in os.listdir(data_dir):
        item_path = os.path.join(data_dir, item)
        if os.path.isdir(item_path):
            # Check if folder name is a year
            try:
                year = int(item)
                if 1990 <= year <= 2030:
                    years.append((year, item_path))
            except ValueError:
                continue
    
    years.sort(key=lambda x: x[0])
    return years


def load_year_data(year_dir, year):
    """Load data for a single year"""
    
    # Find the main lm_data file
    patterns = [
        os.path.join(year_dir, 'lm_data*.txt'),
        os.path.join(year_dir, 'LM_DATA*.txt'),
        os.path.join(year_dir, '*lm_data*.txt'),
    ]
    
    filepath = None
    for pattern in patterns:
        matches = glob.glob(pattern)
        if matches:
            filepath = matches[0]
            break
    
    if not filepath:
        return None
    
    try:
        df = pd.read_csv(
            filepath,
            sep='|',
            low_memory=False,
            encoding='latin-1',
            on_bad_lines='skip'
        )
        df['YEAR'] = year
        return df
    except Exception as e:
        print(f"  Error loading {year}: {e}")
        return None


def load_officer_data(year_dir, year):
    """Load officer compensation data for a single year"""
    
    patterns = [
        os.path.join(year_dir, '*disbursements_emp_off*.txt'),
        os.path.join(year_dir, '*EMP_OFF*.txt'),
        os.path.join(year_dir, 'ar_disbursements_emp_off*.txt'),
    ]
    
    filepath = None
    for pattern in patterns:
        matches = glob.glob(pattern)
        if matches:
            filepath = matches[0]
            break
    
    if not filepath:
        return None
    
    try:
        df = pd.read_csv(
            filepath,
            sep='|',
            low_memory=False,
            encoding='latin-1',
            on_bad_lines='skip'
        )
        df['YEAR'] = year
        return df
    except:
        return None


def classify_union(row):
    """Classify union to major international category"""
    check_fields = []
    for field in ['UNION_NAME', 'AFFIL_NAME', 'DESIG_NAME']:
        if field in row.index and pd.notna(row[field]):
            check_fields.append(str(row[field]).lower())
    
    combined_text = ' '.join(check_fields)
    
    for union, keywords in MAJOR_UNIONS.items():
        for keyword in keywords:
            if keyword in combined_text:
                return union
    
    union_indicators = ['local', 'district council', 'joint board', 'afl-cio', 'union']
    for indicator in union_indicators:
        if indicator in combined_text:
            return 'Other Union'
    
    return None


def standardize_columns(df):
    """Standardize column names across different year formats"""
    col_map = {}
    for col in df.columns:
        col_lower = col.lower()
        if 'union_name' in col_lower:
            col_map[col] = 'UNION_NAME'
        elif 'f_num' in col_lower:
            col_map[col] = 'F_NUM'
        elif col_lower == 'members' or (col_lower.startswith('member') and 'ship' not in col_lower):
            col_map[col] = 'MEMBERS'
        elif 'total_assets' in col_lower:
            col_map[col] = 'TOTAL_ASSETS'
        elif 'total_receipts' in col_lower:
            col_map[col] = 'TOTAL_RECEIPTS'
        elif 'total_disb' in col_lower:
            col_map[col] = 'TOTAL_DISBURSEMENTS'
        elif 'affil' in col_lower and 'name' in col_lower:
            col_map[col] = 'AFFIL_NAME'
        elif 'desig' in col_lower and 'name' in col_lower:
            col_map[col] = 'DESIG_NAME'
        elif 'form_type' in col_lower or col_lower == 'form_type':
            col_map[col] = 'FORM_TYPE'
        elif col_lower == 'rpt_id':
            col_map[col] = 'RPT_ID'
    
    return df.rename(columns=col_map)


def standardize_officer_columns(df):
    """Standardize officer data columns"""
    col_map = {}
    for col in df.columns:
        col_lower = col.lower()
        if 'rpt_id' in col_lower:
            col_map[col] = 'RPT_ID'
        elif 'last_name' in col_lower or col_lower == 'lname':
            col_map[col] = 'LAST_NAME'
        elif 'first_name' in col_lower or col_lower == 'fname':
            col_map[col] = 'FIRST_NAME'
        elif 'title' in col_lower:
            col_map[col] = 'TITLE'
        elif 'gross_sal' in col_lower:
            col_map[col] = 'GROSS_SALARY'
        elif col_lower == 'total' or 'total_comp' in col_lower:
            col_map[col] = 'TOTAL_COMPENSATION'
        elif 'emp_off' in col_lower or 'empoff' in col_lower:
            col_map[col] = 'EMP_OFF_CODE'
    
    return df.rename(columns=col_map)


def create_excel_report(all_data, yearly_summary, union_trends, officer_trends, output_path):
    """Create comprehensive Excel workbook with multiple analysis sheets"""
    
    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # =========================================================================
    # Sheet 1: Executive Summary
    # =========================================================================
    ws = wb.active
    ws.title = "Executive Summary"
    
    years_covered = sorted(all_data['YEAR'].unique()) if all_data is not None else []
    
    rows = [
        ["OLMS Union Financial Analysis", ""],
        ["Multi-Year Trend Report", ""],
        ["", ""],
        ["DATA COVERAGE", ""],
        ["Years Analyzed:", f"{min(years_covered)} - {max(years_covered)}" if years_covered else "N/A"],
        ["Total Year-Filings:", f"{len(all_data):,}" if all_data is not None else "0"],
        ["", ""],
    ]
    
    if yearly_summary is not None and not yearly_summary.empty:
        latest = yearly_summary[yearly_summary['YEAR'] == yearly_summary['YEAR'].max()].iloc[0]
        rows.extend([
            ["LATEST YEAR TOTALS", ""],
            ["Year:", int(latest['YEAR'])],
            ["Total Filings:", f"{int(latest['FILING_COUNT']):,}"],
        ])
        if 'MEMBERS' in latest:
            rows.append(["Total Members:", f"{latest['MEMBERS']:,.0f}"])
        if 'TOTAL_ASSETS' in latest:
            rows.append(["Total Assets:", f"${latest['TOTAL_ASSETS']:,.0f}"])
        if 'TOTAL_RECEIPTS' in latest:
            rows.append(["Total Receipts:", f"${latest['TOTAL_RECEIPTS']:,.0f}"])
    
    for row_idx, row_data in enumerate(rows, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx in [1, 2, 4, 8]:
                cell.font = Font(bold=True, size=12)
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    
    # =========================================================================
    # Sheet 2: Yearly Totals
    # =========================================================================
    if yearly_summary is not None and not yearly_summary.empty:
        ws2 = wb.create_sheet("Yearly Totals")
        headers = list(yearly_summary.columns)
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h.replace('_', ' '))
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        
        for row_idx, row in enumerate(yearly_summary.values, 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                h = headers[col_idx - 1]
                if any(x in h for x in ['ASSETS', 'RECEIPTS', 'DISBURSEMENTS']):
                    cell.number_format = MONEY_FORMAT
                elif 'PCT' in h:
                    cell.number_format = PCT_FORMAT
                elif any(x in h for x in ['MEMBERS', 'COUNT', 'YEAR']):
                    cell.number_format = NUM_FORMAT
        
        for col_idx, h in enumerate(headers, 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = max(15, len(h) + 2)
    
    # =========================================================================
    # Sheet 3: Union Trends (Members over time)
    # =========================================================================
    if union_trends is not None and not union_trends.empty:
        ws3 = wb.create_sheet("Union Member Trends")
        headers = list(union_trends.columns)
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col_idx, value=str(h))
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        
        for row_idx, row in enumerate(union_trends.values, 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx > 1:  # Skip union name column
                    cell.number_format = NUM_FORMAT
        
        ws3.column_dimensions['A'].width = 20
        for col_idx in range(2, len(headers) + 1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = 12
    
    # =========================================================================
    # Sheet 4: Union Trends (Assets over time)
    # =========================================================================
    if 'assets_trends' in dir() and assets_trends is not None:
        ws4 = wb.create_sheet("Union Asset Trends")
        # Similar structure...
    
    # =========================================================================
    # Sheet 5: Top 20 Unions (Latest Year)
    # =========================================================================
    if all_data is not None and not all_data.empty:
        ws5 = wb.create_sheet("Top 20 Latest Year")
        
        latest_year = all_data['YEAR'].max()
        latest_data = all_data[all_data['YEAR'] == latest_year].copy()
        
        # Aggregate by union category
        agg_cols = {}
        for col in ['MEMBERS', 'TOTAL_ASSETS', 'TOTAL_RECEIPTS', 'TOTAL_DISBURSEMENTS']:
            if col in latest_data.columns:
                agg_cols[col] = 'sum'
        agg_cols['UNION_NAME'] = 'count'
        
        if 'UNION_CATEGORY' in latest_data.columns:
            top_unions = latest_data.groupby('UNION_CATEGORY').agg(agg_cols).reset_index()
            top_unions = top_unions.rename(columns={'UNION_NAME': 'FILING_COUNT'})
            if 'MEMBERS' in top_unions.columns:
                top_unions = top_unions.sort_values('MEMBERS', ascending=False).head(20)
            
            headers = list(top_unions.columns)
            for col_idx, h in enumerate(headers, 1):
                cell = ws5.cell(row=1, column=col_idx, value=h.replace('_', ' '))
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            
            for row_idx, row in enumerate(top_unions.values, 2):
                for col_idx, val in enumerate(row, 1):
                    cell = ws5.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    h = headers[col_idx - 1]
                    if any(x in h for x in ['ASSETS', 'RECEIPTS', 'DISBURSEMENTS']):
                        cell.number_format = MONEY_FORMAT
                    elif any(x in h for x in ['MEMBERS', 'COUNT']):
                        cell.number_format = NUM_FORMAT
            
            for col_idx, h in enumerate(headers, 1):
                ws5.column_dimensions[get_column_letter(col_idx)].width = max(18, len(h) + 2)
    
    # =========================================================================
    # Sheet 6: Officer Compensation Trends
    # =========================================================================
    if officer_trends is not None and not officer_trends.empty:
        ws6 = wb.create_sheet("Officer Comp Trends")
        headers = list(officer_trends.columns)
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws6.cell(row=1, column=col_idx, value=str(h).replace('_', ' '))
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        
        for row_idx, row in enumerate(officer_trends.head(100).values, 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws6.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx > 1:
                    cell.number_format = MONEY_FORMAT
        
        ws6.column_dimensions['A'].width = 25
        for col_idx in range(2, len(headers) + 1):
            ws6.column_dimensions[get_column_letter(col_idx)].width = 12
    
    # =========================================================================
    # Sheet 7: All Data (for researchers)
    # =========================================================================
    if all_data is not None and not all_data.empty:
        ws7 = wb.create_sheet("All Union Data")
        
        export_cols = [c for c in ['YEAR', 'UNION_CATEGORY', 'UNION_NAME', 'F_NUM', 
                                   'MEMBERS', 'TOTAL_ASSETS', 'TOTAL_RECEIPTS', 
                                   'TOTAL_DISBURSEMENTS', 'AFFIL_NAME']
                      if c in all_data.columns]
        
        export_data = all_data[export_cols].sort_values(['YEAR', 'MEMBERS'], ascending=[True, False])
        
        for col_idx, h in enumerate(export_cols, 1):
            cell = ws7.cell(row=1, column=col_idx, value=h.replace('_', ' '))
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        
        # Limit to 50000 rows for Excel performance
        for row_idx, row in enumerate(export_data.head(50000).values, 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws7.cell(row=row_idx, column=col_idx, value=val)
                h = export_cols[col_idx - 1]
                if any(x in h for x in ['ASSETS', 'RECEIPTS', 'DISBURSEMENTS']):
                    cell.number_format = MONEY_FORMAT
                elif any(x in h for x in ['MEMBERS', 'YEAR']):
                    cell.number_format = NUM_FORMAT
        
        for col_idx, h in enumerate(export_cols, 1):
            ws7.column_dimensions[get_column_letter(col_idx)].width = 18
    
    wb.save(output_path)
    print(f"\n*** Report saved to: {output_path} ***")


def main():
    parser = argparse.ArgumentParser(description='Multi-year OLMS union analysis')
    parser.add_argument('--data-dir', required=True, help='Parent folder containing year subfolders')
    parser.add_argument('--output', default='union_trends.xlsx', help='Output Excel file')
    
    args = parser.parse_args()
    
    print("=" * 70)
    print("OLMS Multi-Year Union Financial Analysis (2000-2025)")
    print("=" * 70)
    print(f"Data folder: {args.data_dir}")
    print(f"Output file: {args.output}")
    print("=" * 70)
    
    # Find year folders
    print("\nScanning for year folders...")
    year_folders = find_year_folders(args.data_dir)
    print(f"Found {len(year_folders)} year folders: {[y[0] for y in year_folders]}")
    
    # Load all years
    print("\n" + "=" * 70)
    print("Loading data from each year...")
    print("=" * 70)
    
    all_dfs = []
    all_officer_dfs = []
    
    for year, year_path in year_folders:
        print(f"\n{year}:", end=" ")
        
        df = load_year_data(year_path, year)
        if df is not None:
            df = standardize_columns(df)
            
            # Convert numeric columns
            for col in ['MEMBERS', 'TOTAL_ASSETS', 'TOTAL_RECEIPTS', 'TOTAL_DISBURSEMENTS']:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Filter to LM-2 if possible
            if 'FORM_TYPE' in df.columns:
                df = df[df['FORM_TYPE'].str.contains('LM-2', case=False, na=False)]
            
            # Classify unions
            df['UNION_CATEGORY'] = df.apply(classify_union, axis=1)
            df = df[df['UNION_CATEGORY'].notna()]
            
            print(f"{len(df):,} union records")
            all_dfs.append(df)
        else:
            print("No data found")
        
        # Load officer data
        off_df = load_officer_data(year_path, year)
        if off_df is not None:
            off_df = standardize_officer_columns(off_df)
            for col in ['GROSS_SALARY', 'TOTAL_COMPENSATION']:
                if col in off_df.columns:
                    off_df[col] = pd.to_numeric(off_df[col], errors='coerce')
            all_officer_dfs.append(off_df)
    
    if not all_dfs:
        print("\nERROR: No data loaded from any year!")
        return
    
    # Combine all years
    print("\n" + "=" * 70)
    print("Combining all years...")
    all_data = pd.concat(all_dfs, ignore_index=True)
    print(f"Total records: {len(all_data):,}")
    
    # Create yearly summary
    print("\nCreating yearly summary...")
    agg_cols = {'UNION_NAME': 'count'}
    for col in ['MEMBERS', 'TOTAL_ASSETS', 'TOTAL_RECEIPTS', 'TOTAL_DISBURSEMENTS']:
        if col in all_data.columns:
            agg_cols[col] = 'sum'
    
    yearly_summary = all_data.groupby('YEAR').agg(agg_cols).reset_index()
    yearly_summary = yearly_summary.rename(columns={'UNION_NAME': 'FILING_COUNT'})
    yearly_summary = yearly_summary.sort_values('YEAR')
    
    print("\nYEARLY TOTALS:")
    print(yearly_summary.to_string(index=False))
    
    # Create union trends (pivot table: unions x years)
    print("\nCreating union membership trends...")
    if 'MEMBERS' in all_data.columns:
        union_year_members = all_data.groupby(['UNION_CATEGORY', 'YEAR'])['MEMBERS'].sum().reset_index()
        union_trends = union_year_members.pivot(index='UNION_CATEGORY', columns='YEAR', values='MEMBERS')
        union_trends = union_trends.reset_index()
        
        # Sort by most recent year
        latest_year = max(union_trends.columns[1:])
        union_trends = union_trends.sort_values(latest_year, ascending=False)
    else:
        union_trends = pd.DataFrame()
    
    # Officer compensation trends
    print("\nCreating officer compensation trends...")
    officer_trends = pd.DataFrame()
    if all_officer_dfs:
        all_officers = pd.concat(all_officer_dfs, ignore_index=True)
        
        # Get average/max compensation by year
        if 'TOTAL_COMPENSATION' in all_officers.columns:
            comp_col = 'TOTAL_COMPENSATION'
        elif 'GROSS_SALARY' in all_officers.columns:
            comp_col = 'GROSS_SALARY'
        else:
            comp_col = None
        
        if comp_col:
            # Top paid by year
            officer_by_year = all_officers.groupby('YEAR')[comp_col].agg(['mean', 'max', 'median']).reset_index()
            officer_by_year.columns = ['YEAR', 'AVG_COMPENSATION', 'MAX_COMPENSATION', 'MEDIAN_COMPENSATION']
            officer_trends = officer_by_year
    
    # Create report
    print("\n" + "=" * 70)
    print("Creating Excel report...")
    print("=" * 70)
    
    create_excel_report(all_data, yearly_summary, union_trends, officer_trends, args.output)
    
    print("\n" + "=" * 70)
    print("ANALYSIS COMPLETE!")
    print("=" * 70)
    print(f"\nYears covered: {min(all_data['YEAR'])} - {max(all_data['YEAR'])}")
    print(f"Total records: {len(all_data):,}")
    print(f"Output file: {args.output}")


if __name__ == "__main__":
    main()
