"""Build an Excel workbook so Jacob can pick a --min-score threshold
for the F7 orphan rematch executor (B.3.x Option C).

Inputs:
  files/orphan_rematch_2026_05_08/best_matches.csv
  files/orphan_rematch_2026_05_08/review_queue.csv

Output:
  files/orphan_rematch_2026_05_08/Orphan_Rematch_Review.xlsx

Sheets:
  1. Threshold Tradeoffs   — table showing what happens at each
                              candidate --min-score (1.00, 0.98, 0.96,
                              0.92). The "decide here" sheet.
  2. Score x Tier Matrix   — counts pivoted by score x rule_engine_tier
  3. Score x Source Matrix — counts pivoted by score x source
  4. All Best Matches      — the full 1,179 best-match rows; sortable
  5. Sample by Bucket      — 12 sample rows per (score, source) pair
                              for eyeball quality check
  6. Vetoed (Caught FPs)   — re-runs the SQL + rule engine to surface
                              the 13 series-fragment / person-name pairs
                              the rule engine blocked
  7. Review Queue Sample   — first 200 rows of the 4,895 tier_C queue
                              (matches kept in main pool but flagged
                              for human spot-check)

Invoke:
  py scripts/analysis/build_orphan_rematch_review_xlsx.py
"""
from __future__ import annotations

import csv
import sys
import random
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))

# Re-use the executor's helpers so the vetoed-pair sheet is derived
# from the same code path that the rematch script uses.
from scripts.matching.rematch_recoverable_orphans import (  # noqa: E402
    SOURCES,
    TIERS,
    attempt_match,
    classify_with_rule_engine,
)
from db_config import get_connection  # noqa: E402


OUT_DIR = REPO / "files" / "orphan_rematch_2026_05_08"
BEST_CSV = OUT_DIR / "best_matches.csv"
REVIEW_CSV = OUT_DIR / "review_queue.csv"
XLSX_OUT = OUT_DIR / "Orphan_Rematch_Review.xlsx"


# ----------------------------------------------------------------------
# Styling helpers
# ----------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
ZEBRA_FILL  = PatternFill("solid", fgColor="F2F2F2")
GOOD_FILL   = PatternFill("solid", fgColor="C6EFCE")  # green
WARN_FILL   = PatternFill("solid", fgColor="FFEB9C")  # amber
BAD_FILL    = PatternFill("solid", fgColor="FFC7CE")  # pink


def _write_headers(ws, headers, row=1):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


def _autosize(ws, sample_rows=200):
    """Approximate column auto-fit. openpyxl doesn't have a built-in
    auto-fit so we measure the longest string in the first N rows."""
    for col_idx, col in enumerate(ws.iter_cols(min_row=1,
                                                max_row=sample_rows + 1),
                                  start=1):
        max_len = 8
        for cell in col:
            if cell.value is None:
                continue
            v = str(cell.value)
            if len(v) > max_len:
                max_len = min(len(v), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


# ----------------------------------------------------------------------
# Data loaders
# ----------------------------------------------------------------------

def load_best_matches() -> list[dict]:
    if not BEST_CSV.exists():
        sys.exit(f"Missing input: {BEST_CSV}\n"
                 "Run rematch_recoverable_orphans.py with --out-csv first.")
    with open(BEST_CSV, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_review_queue() -> list[dict]:
    if not REVIEW_CSV.exists():
        return []
    with open(REVIEW_CSV, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def derive_vetoed_pairs() -> list[dict]:
    """Re-run the SQL + rule engine to capture the candidates the
    engine vetoed (tier_series_demoted). These don't make it to the
    best-match CSV by design; this sheet shows what was caught."""
    conn = get_connection()
    cur = conn.cursor()
    vetoes: list[dict] = []
    for source_key, _table, name_col, state_col, zip_col, display_col, city_col, ein_col in SOURCES:
        for tier in TIERS:
            matches = attempt_match(
                cur, source_key, name_col, state_col, zip_col, tier, None,
                display_col=display_col, city_col=city_col, ein_col=ein_col,
            )
            for m in matches:
                classify_with_rule_engine(m)
                if m.get("rule_engine_tier") == "tier_series_demoted":
                    vetoes.append(m)
    cur.close()
    conn.close()
    # Dedup by (f7_employer_id, source, source_id)
    seen = set()
    uniq = []
    for v in vetoes:
        key = (v["f7_employer_id"], v["source"], v["source_id"])
        if key in seen:
            continue
        seen.add(key)
        uniq.append(v)
    return uniq


# ----------------------------------------------------------------------
# Sheet builders
# ----------------------------------------------------------------------

def build_threshold_sheet(wb: Workbook, rows: list[dict]):
    """The "decide here" sheet. Rows are min-score thresholds; columns
    are cumulative cohort size + fraction rule-engine-confirmed +
    per-source breakdown + recommended action."""
    ws = wb.create_sheet("Threshold Tradeoffs", index=0)

    intro = [
        ("F7 ORPHAN REMATCH — THRESHOLD DECISION SHEET", True),
        ("", False),
        ("This is the sheet to use when picking --min-score for the commit run.", False),
        ("Each row shows what happens at a different threshold.", False),
        ("", False),
        ("Score legend:", True),
        ("  1.00 = NAME_STANDARD_STATE_ZIP_EXACT  (strongest SQL match)", False),
        ("  0.98 = NAME_STANDARD_STATE_EXACT", False),
        ("  0.96 = NAME_AGGRESSIVE upgraded by the rule engine (tier_A confirmation)", False),
        ("  0.92 = NAME_AGGRESSIVE_STATE_EXACT, NO rule engine confirmation", False),
        ("", False),
        ("Rule engine tier legend:", True),
        ("  tier_A_auto_merge = >=96% precision (Haiku-validated). Safe.", False),
        ("  tier_C_review     = 50-90% precision. Kept in pool but flagged.", False),
        ("  tier_D_different  = no rule fired but SQL match strong. Lower confidence.", False),
        ("", False),
    ]
    r = 1
    for line, bold in intro:
        cell = ws.cell(row=r, column=1, value=line)
        if bold:
            cell.font = Font(bold=True, size=12 if r == 1 else 11)
        r += 1

    headers = [
        "min_score (--min-score)",
        "Cohort size (matches written)",
        "% of full pool",
        "Rule-engine confirmed (tier_A)",
        "Tier_C review (kept but flagged)",
        "Tier_D no-rule (kept anyway)",
        "OSHA",
        "WHD",
        "990",
        "SAM",
        "Recommendation",
    ]
    header_row = r
    _write_headers(ws, headers, row=header_row)
    r = header_row + 1

    thresholds = [1.00, 0.98, 0.96, 0.92]
    full_pool = len(rows)
    recommendations = {
        1.00: "Most conservative. Only ZIP-confirmed exact name matches. ~100% precision.",
        0.98: "Safe. NAME_STANDARD exact matches across state. ~98% precision.",
        0.96: "RECOMMENDED. Rule-engine corroboration lifts NAME_AGGRESSIVE matches that pass H1+H16/H2+H3. ~96% precision.",
        0.92: "Aggressive. Includes 0.92 NAME_AGGRESSIVE matches with NO rule engine confirmation. ~92% precision per V2 cascade definition.",
    }

    for thresh in thresholds:
        cohort = [m for m in rows if float(m["score"]) >= thresh]
        n = len(cohort)
        n_a = sum(1 for m in cohort if m.get("rule_engine_tier") == "tier_A_auto_merge")
        n_c = sum(1 for m in cohort if m.get("rule_engine_tier") == "tier_C_review")
        n_d = sum(1 for m in cohort if m.get("rule_engine_tier") == "tier_D_different")
        by_src = {"osha": 0, "whd": 0, "990": 0, "sam": 0}
        for m in cohort:
            by_src[m["source"]] += 1

        row = [
            thresh,
            n,
            round(100 * n / full_pool, 1),
            n_a,
            n_c,
            n_d,
            by_src["osha"],
            by_src["whd"],
            by_src["990"],
            by_src["sam"],
            recommendations[thresh],
        ]
        for col, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            if thresh == 0.96 and col == 11:  # highlight the recommendation
                cell.fill = GOOD_FILL
                cell.font = Font(bold=True)
            elif thresh == 1.00:
                if col == 1:
                    cell.fill = WARN_FILL  # too conservative — cohort small
            elif thresh == 0.92 and col >= 4 and col <= 6:
                # the unconfirmed bucket (tier_C / tier_D) — these are
                # what tightens up if you raise the threshold to 0.96
                if v and v > 0:
                    cell.fill = WARN_FILL
        r += 1

    # Footer note
    r += 1
    ws.cell(row=r, column=1, value=(
        "Note: 'Rule-engine confirmed (tier_A)' counts pairs where H1, H2, H6, H11, "
        "H13, H15, or H16 fired (validated >=96% precision against 31,532 Haiku-labeled "
        "pairs). At threshold=0.96 you get all rule-engine-confirmed matches plus all "
        "NAME_STANDARD exact matches; at 0.98 you DROP the 126 rule-engine upgrades."
    )).font = Font(italic=True, size=10)
    r += 2
    ws.cell(row=r, column=1, value=(
        "13 candidates were VETOED (rule_engine tier_series_demoted) and never "
        "appeared in the best-match pool — see the 'Vetoed (Caught FPs)' sheet for "
        "what was caught. Those are series-fragment matches like "
        "'XYZ HOLDINGS SERIES 1' vs 'XYZ HOLDINGS SERIES 2' and reversed person "
        "names like 'WILLIAMS JAMES K' vs 'WILLIAMS JAMES P'."
    )).font = Font(italic=True, size=10)

    _autosize(ws)


def build_score_tier_matrix(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Score x Tier Matrix")
    tiers = ["tier_A_auto_merge", "tier_B_high_conf", "tier_C_review", "tier_D_different"]
    matrix: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for m in rows:
        s = f"{float(m['score']):.2f}"
        t = m.get("rule_engine_tier") or "(no tag)"
        matrix[s][t] += 1

    headers = ["score"] + tiers + ["Total"]
    _write_headers(ws, headers)
    r = 2
    for s in sorted(matrix.keys(), reverse=True):
        ws.cell(row=r, column=1, value=s).font = Font(bold=True)
        total = 0
        for col, t in enumerate(tiers, start=2):
            v = matrix[s].get(t, 0)
            ws.cell(row=r, column=col, value=v)
            total += v
        ws.cell(row=r, column=len(tiers) + 2, value=total).font = Font(bold=True)
        r += 1

    ws.cell(row=r + 1, column=1, value=(
        "How to read: row 0.96 column tier_A_auto_merge = the 126 rule-engine "
        "upgrades. These are NAME_AGGRESSIVE matches that the engine confirmed "
        "as tier_A precision."
    )).font = Font(italic=True, size=10)
    _autosize(ws)


def build_score_source_matrix(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Score x Source Matrix")
    sources = ["osha", "whd", "990", "sam"]
    matrix: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for m in rows:
        s = f"{float(m['score']):.2f}"
        src = m["source"]
        matrix[s][src] += 1

    headers = ["score"] + sources + ["Total"]
    _write_headers(ws, headers)
    r = 2
    for s in sorted(matrix.keys(), reverse=True):
        ws.cell(row=r, column=1, value=s).font = Font(bold=True)
        total = 0
        for col, src in enumerate(sources, start=2):
            v = matrix[s].get(src, 0)
            ws.cell(row=r, column=col, value=v)
            total += v
        ws.cell(row=r, column=len(sources) + 2, value=total).font = Font(bold=True)
        r += 1
    _autosize(ws)


def build_all_best_matches(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("All Best Matches")
    headers = list(rows[0].keys()) if rows else []
    _write_headers(ws, headers)
    for r_idx, m in enumerate(rows, start=2):
        for c_idx, h in enumerate(headers, start=1):
            v = m.get(h, "")
            # Try numeric where it makes sense
            if h == "score":
                try:
                    v = float(v)
                except (TypeError, ValueError):
                    pass
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            if r_idx % 2 == 0:
                cell.fill = ZEBRA_FILL
    ws.freeze_panes = "A2"
    _autosize(ws, sample_rows=50)


def build_sample_by_bucket(wb: Workbook, rows: list[dict], n_per_bucket=12):
    """For each (score, source) pair, pick up to n_per_bucket samples
    so Jacob can eyeball name quality."""
    ws = wb.create_sheet("Sample by Bucket")
    by_bucket: dict[tuple, list[dict]] = defaultdict(list)
    for m in rows:
        key = (f"{float(m['score']):.2f}", m["source"])
        by_bucket[key].append(m)

    rng = random.Random(42)  # deterministic samples
    for k in by_bucket:
        rng.shuffle(by_bucket[k])
        by_bucket[k] = by_bucket[k][:n_per_bucket]

    headers = [
        "score", "source", "method", "rule_engine_tier", "rule_engine_rule",
        "f7_name", "source_name_norm", "f7_state",
    ]
    _write_headers(ws, headers)
    r = 2
    for (score, source) in sorted(by_bucket.keys(), key=lambda x: (-float(x[0]), x[1])):
        # bucket header bar
        ws.cell(row=r, column=1, value=f"--- score={score} source={source} ({len(by_bucket[(score, source)])} samples) ---").font = Font(bold=True, italic=True)
        ws.cell(row=r, column=1).fill = WARN_FILL
        r += 1
        for m in by_bucket[(score, source)]:
            for col, h in enumerate(headers, start=1):
                ws.cell(row=r, column=col, value=m.get(h, ""))
            r += 1
        r += 1
    ws.freeze_panes = "A2"
    _autosize(ws, sample_rows=80)


def build_vetoed_sheet(wb: Workbook, vetoes: list[dict]):
    ws = wb.create_sheet("Vetoed (Caught FPs)")
    intro = [
        ("CANDIDATES THE RULE ENGINE BLOCKED (would-be false positives)", True),
        ("", False),
        ("These pairs PASSED the SQL exact-match cascade (NAME_STANDARD or NAME_AGGRESSIVE)", False),
        ("but the rule engine vetoed them — H4 catches series fragments, person_name_block", False),
        ("catches reversed-name pairs like 'WILLIAMS JAMES K' vs 'WILLIAMS JAMES P'.", False),
        ("", False),
        ("In the 2026-05-06 dry-run (before this integration), all of these would have", False),
        ("been written to unified_match_log as active matches. Now they're blocked.", False),
        ("", False),
    ]
    r = 1
    for line, bold in intro:
        cell = ws.cell(row=r, column=1, value=line)
        if bold:
            cell.font = Font(bold=True, size=12 if r == 1 else 11)
        r += 1

    headers = [
        "rule (rule_engine_rule)", "source", "method",
        "f7_name", "source_name_norm",
        "f7_state", "f7_zip", "source_zip",
    ]
    _write_headers(ws, headers, row=r)
    r += 1
    for m in vetoes:
        for col, h in enumerate(headers, start=1):
            key = h.split(" (")[0] if "(" in h else h
            if h.startswith("rule "):
                key = "rule_engine_rule"
            ws.cell(row=r, column=col, value=m.get(key, ""))
        r += 1
    _autosize(ws)


def build_review_queue_sample(wb: Workbook, queue_rows: list[dict], n=200):
    ws = wb.create_sheet("Review Queue Sample")
    intro = [
        ("REVIEW QUEUE SAMPLE (first 200 of 4,895 rows)", True),
        ("", False),
        ("These are tier_C_review matches: SQL match passed the exact-name floor", False),
        ("but only weak rule-engine corroboration fired (e.g., H1 alone). They're", False),
        ("KEPT in the main pool (will be written if --min-score lets them in) but", False),
        ("ALSO flagged for human review.", False),
        ("", False),
        ("Full file: files/orphan_rematch_2026_05_08/review_queue.csv", False),
        ("", False),
    ]
    r = 1
    for line, bold in intro:
        cell = ws.cell(row=r, column=1, value=line)
        if bold:
            cell.font = Font(bold=True, size=12 if r == 1 else 11)
        r += 1

    if not queue_rows:
        ws.cell(row=r, column=1, value="(no review-queue CSV present)")
        return

    headers = [k for k in queue_rows[0].keys()]
    _write_headers(ws, headers, row=r)
    r += 1
    for m in queue_rows[:n]:
        for col, h in enumerate(headers, start=1):
            v = m.get(h, "")
            if h == "score":
                try:
                    v = float(v)
                except (TypeError, ValueError):
                    pass
            ws.cell(row=r, column=col, value=v)
        r += 1
    ws.freeze_panes = f"A{r - n}" if r > n else "A2"
    _autosize(ws, sample_rows=50)


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------

def main():
    print(f"Loading inputs from {OUT_DIR}/")
    best = load_best_matches()
    review = load_review_queue()
    print(f"  best_matches.csv: {len(best):,} rows")
    print(f"  review_queue.csv: {len(review):,} rows")

    print("Deriving vetoed candidates (re-running SQL + rule engine)...")
    vetoes = derive_vetoed_pairs()
    print(f"  {len(vetoes):,} vetoed pairs found")

    print("Building workbook...")
    wb = Workbook()
    # Remove the default sheet
    default = wb.active
    wb.remove(default)

    build_threshold_sheet(wb, best)
    build_score_tier_matrix(wb, best)
    build_score_source_matrix(wb, best)
    build_all_best_matches(wb, best)
    build_sample_by_bucket(wb, best)
    build_vetoed_sheet(wb, vetoes)
    build_review_queue_sample(wb, review)

    XLSX_OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_OUT)
    print(f"\nWrote: {XLSX_OUT}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
